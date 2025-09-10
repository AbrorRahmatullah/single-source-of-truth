import csv
from decimal import Decimal, InvalidOperation
import io
import math
import pandas as pd
import pyodbc
import os
import re
import logging
import openpyxl

from flask import Flask, flash, make_response, request, render_template, jsonify, redirect, session, url_for, send_file
from functools import wraps
from waitress import serve
from werkzeug.utils import secure_filename
from datetime import datetime, time, date, timedelta
from flask_bcrypt import Bcrypt
from app.config import get_db_connection
from io import BytesIO

app = Flask(__name__)
bcrypt = Bcrypt(app)

app.config['UPLOAD_FOLDER'] = 'uploads'
app.secret_key = 'rahasiayangsangatrahasia'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Set idle session lifetime to 30 minutes
IDLE_TIMEOUT = timedelta(minutes=30)

# Pastikan folder upload ada
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Konfigurasi logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.before_request
def check_idle_timeout():
    # Abaikan pengecekan di halaman login atau static file
    if request.endpoint in ['login', 'static']:
        return

    last_activity = session.get("last_activity")
    if last_activity:
        last_activity_dt = datetime.fromisoformat(last_activity)  # konversi dari string ke datetime
        if datetime.now() - last_activity_dt > IDLE_TIMEOUT:
            session.clear()
            return redirect(url_for("login"))

    # Update waktu terakhir aktivitas
    session["last_activity"] = datetime.now().isoformat()

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ['xlsx', 'xls']

def normalize_column_name(col_name):
    """Normalize column name for comparison"""
    if pd.isna(col_name):
        return ""
    return str(col_name).strip()

def find_matching_column(excel_header, db_columns, strict_mode=False):
    """
    Find matching database column for Excel header with fallback strategies
    
    Args:
        excel_header: Header from Excel file
        db_columns: List of database column names
        strict_mode: If True, only exact matches allowed
    
    Returns:
        matched_db_column or None
    """
    if not excel_header or pd.isna(excel_header):
        return None
    
    excel_norm = normalize_column_name(excel_header)
    if not excel_norm:
        return None
    
    # Strategy 1: Exact match (normalized)
    for db_col in db_columns:
        db_norm = normalize_column_name(db_col)
        if excel_norm == db_norm:
            return db_col
    
    if strict_mode:
        return None
    
    # Strategy 2: Space to underscore conversion
    excel_underscore = excel_norm.replace(' ', '_')
    for db_col in db_columns:
        db_norm = normalize_column_name(db_col)
        db_underscore = db_norm.replace(' ', '_')
        if excel_underscore == db_underscore:
            return db_col
    
    # Strategy 3: Remove all separators (spaces, underscores, hyphens)
    excel_clean = re.sub(r'[\s_-]+', '', excel_norm)
    if len(excel_clean) >= 3:  # Only for meaningful length
        for db_col in db_columns:
            db_norm = normalize_column_name(db_col)
            db_clean = re.sub(r'[\s_-]+', '', db_norm)
            if excel_clean == db_clean:
                return db_col
    
    # Strategy 4: Substring matching (minimum 4 characters)
    if len(excel_norm) >= 4:
        for db_col in db_columns:
            db_norm = normalize_column_name(db_col)
            if len(db_norm) >= 4:
                if excel_norm in db_norm or db_norm in excel_norm:
                    return db_col
    
    return None

def find_primary_header_row(df, primary_header_pattern=None):
    def normalize_text(text):
        if pd.isna(text):
            return ""
        return str(text).lower().strip()

    def is_likely_header_row(row_values):
        non_empty_count = 0
        text_count = 0
        numeric_count = 0

        for val in row_values:
            if pd.notna(val) and str(val).strip():
                non_empty_count += 1
                try:
                    float(str(val))
                    numeric_count += 1
                except ValueError:
                    text_count += 1

        if non_empty_count < 2:
            return False

        if numeric_count == non_empty_count:
            return False

        return True

    def check_column_pattern(row_values):
        text_values = [normalize_text(val) for val in row_values if pd.notna(val) and str(val).strip()]
        column_pattern_count = sum(1 for val in text_values if val.startswith('column_'))
        if column_pattern_count >= 2:
            return 10

        header_indicators = ['col', 'field', 'data', 'info', 'value', 'item']
        score = 0
        for val in text_values:
            for indicator in header_indicators:
                if indicator in val:
                    score += 1
        return score

    # Step 1: jika diberikan primary_header_pattern
    if primary_header_pattern:
        pattern_normalized = normalize_text(primary_header_pattern)
        for idx, row in df.iterrows():
            row_values = [normalize_text(val) for val in row]
            for val in row_values:
                if pattern_normalized in val or val in pattern_normalized:
                    if is_likely_header_row(row.values):
                        logger.info(f"Header row ditemukan di baris {idx + 1} berdasarkan pattern '{primary_header_pattern}'")
                        return idx, primary_header_pattern

    # Step 2: auto-detection
    best_header_row = None
    best_score = -1
    detected_primary = None

    common_header_keywords = [
        'number', 'name', 'id', 'code', 'facility', 'location', 'type', 
        'date', 'status', 'description', 'value', 'amount', 'quantity',
        'column', 'field', 'data', 'info'
    ]

    for idx, row in df.iterrows():
        if idx > 20:
            break

        if not is_likely_header_row(row.values):
            continue

        score = 0
        primary_candidate = None
        row_values = [str(val) for val in row if pd.notna(val)]

        # Penambahan: beri preferensi lebih tinggi jika ini baris pertama
        if idx == 0:
            score += 5  # penalti negatif untuk baris bukan pertama

        # Skor berdasarkan keyword
        for val in row_values:
            val_norm = normalize_text(val)
            for keyword in common_header_keywords:
                if keyword in val_norm:
                    score += 1
                    if not primary_candidate and (
                        'number' in val_norm or 'id' in val_norm or 'column' in val_norm
                    ):
                        primary_candidate = str(val).strip()

        # Skor berdasarkan pola column_
        score += check_column_pattern(row.values)

        # Tambahan bonus untuk jumlah kolom
        non_empty_count = sum(1 for val in row_values if str(val).strip())
        if non_empty_count >= 3:
            score += 2

        if score > best_score:
            best_score = score
            best_header_row = idx
            detected_primary = primary_candidate or str(row_values[0]).strip()

    if best_header_row is not None:
        logger.info(f"Header row auto-detected di baris {best_header_row + 1}, primary header: '{detected_primary}', score: {best_score}")
        return best_header_row, detected_primary

    # Fallback
    logger.warning("Menggunakan fallback detection...")
    for idx, row in df.iterrows():
        if idx > 10:
            break
        if is_likely_header_row(row.values):
            row_values = [str(val) for val in row if pd.notna(val)]
            if row_values:
                detected_primary = str(row_values[0]).strip()
                logger.info(f"Header row fallback detected di baris {idx + 1}, primary header: '{detected_primary}'")
                return idx, detected_primary

    raise ValueError("Tidak dapat menemukan baris header. Pastikan file Excel memiliki baris header yang jelas.")

def find_header_row_and_validate(df, required_headers, primary_header_pattern=None):
    """
    Enhanced header finding with strict column validation
    """
    def is_likely_header_row(row_data, required_headers):
        """Check if a row looks like a header row"""
        non_empty_count = sum(1 for val in row_data if pd.notna(val) and str(val).strip())
        if non_empty_count < len(required_headers) * 0.5:  # At least 50% filled
            return False
        
        # Check for numeric-only rows (likely data, not headers)
        text_count = 0
        for val in row_data:
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str and not val_str.replace('.', '').replace('-', '').isdigit():
                    text_count += 1
        
        return text_count >= non_empty_count * 0.7  # At least 70% non-numeric

    def normalize_column_name(col_name):
        """Normalize column name for comparison"""
        if pd.isna(col_name):
            return ""
        return str(col_name).strip()

    def convert_spaces_to_underscore(col_name):
        """Convert spaces to underscores in column name"""
        if pd.isna(col_name):
            return ""
        return str(col_name).strip().replace(' ', '_')

    def strict_column_match(excel_headers, required_headers):
        """Strict column matching with enhanced validation"""
        # Remove empty headers
        non_empty_excel_headers = [h for h in excel_headers if h.strip()]
        
        # Validation 1: Check column count
        if len(non_empty_excel_headers) != len(required_headers):
            raise ValueError(
                f"File Excel memiliki {len(non_empty_excel_headers)} kolom, "
                f"tetapi tabel database membutuhkan {len(required_headers)} kolom. "
                f"Excel: {non_empty_excel_headers}, Database: {required_headers}"
            )
        
        # Validation 2: Exact name matching (case-insensitive)
        valid_headers_mapping = []
        found_db_columns = set()
        
        # First attempt: exact matching
        for i, excel_header in enumerate(non_empty_excel_headers):
            normalized_excel = normalize_column_name(excel_header).lower()
            matched_db_col = None
            
            # Find exact match (case-insensitive)
            for db_col in required_headers:
                if db_col not in found_db_columns and db_col.lower() == normalized_excel:
                    matched_db_col = db_col
                    break
            
            if matched_db_col:
                valid_headers_mapping.append((excel_header, matched_db_col))
                found_db_columns.add(matched_db_col)
            else:
                # Column name doesn't match exactly, will try space conversion later
                valid_headers_mapping.append((excel_header, None))
        
        # Check if all columns matched exactly
        unmatched_excel_headers = [mapping[0] for mapping in valid_headers_mapping if mapping[1] is None]
        unmatched_db_headers = [col for col in required_headers if col not in found_db_columns]
        
        if unmatched_excel_headers:
            logger.info(f"Some columns didn't match exactly. Attempting space-to-underscore conversion...")
            
            # Validation 3: Try converting spaces to underscores
            space_converted_mapping = []
            remaining_found_columns = found_db_columns.copy()
            remaining_db_columns = [col for col in required_headers if col not in found_db_columns]
            
            for excel_header, db_match in valid_headers_mapping:
                if db_match is None:  # This column didn't match
                    # Try converting spaces to underscores
                    converted_header = convert_spaces_to_underscore(excel_header).lower()
                    matched_db_col = None
                    
                    # Find match with converted name
                    for db_col in remaining_db_columns:
                        if db_col.lower() == converted_header:
                            matched_db_col = db_col
                            break
                    
                    if matched_db_col:
                        space_converted_mapping.append((excel_header, matched_db_col))
                        remaining_found_columns.add(matched_db_col)
                        remaining_db_columns.remove(matched_db_col)
                        logger.info(f"Successfully converted '{excel_header}' -> '{matched_db_col}' (space to underscore)")
                    else:
                        space_converted_mapping.append((excel_header, None))
                else:
                    space_converted_mapping.append((excel_header, db_match))
            
            # Update the mapping
            valid_headers_mapping = space_converted_mapping
            found_db_columns = remaining_found_columns
        
        # Final validation: Check if all columns are matched
        final_unmatched_excel = [mapping[0] for mapping in valid_headers_mapping if mapping[1] is None]
        final_unmatched_db = [col for col in required_headers if col not in found_db_columns]
        
        if final_unmatched_excel or final_unmatched_db:
            error_msg = "Nama kolom tidak sesuai dengan template database.\n\n"
            
            # if final_unmatched_excel:
            #     error_msg += f"Kolom di Excel yang tidak dikenali: {final_unmatched_excel}\n"
            
            # if final_unmatched_db:
            #     error_msg += f"Kolom database yang tidak ditemukan: {final_unmatched_db}\n"
            
            # error_msg += f"\nKolom yang dibutuhkan database: {required_headers}\n"
            # error_msg += f"Kolom yang ditemukan di Excel: {[h for h in excel_headers if h.strip()]}\n\n"
            error_msg += "Pastikan nama kolom di Excel sesuai dengan template database (case-insensitive). "
            error_msg += "Spasi dalam nama kolom akan otomatis dikonversi ke underscore."
            
            raise ValueError(error_msg)
        
        return valid_headers_mapping, [], []  # No missing headers in strict mode
    
    # Find header row
    header_row = None
    detected_primary = None
    
    # If primary_header_pattern provided, search for it first
    if primary_header_pattern:
        for idx in range(min(20, len(df))):  # Search first 20 rows
            row_data = df.iloc[idx].tolist()
            for val in row_data:
                if pd.notna(val):
                    val_str = str(val).strip().lower()
                    if primary_header_pattern.lower() in val_str:
                        if is_likely_header_row(row_data, required_headers):
                            header_row = idx
                            detected_primary = str(val).strip()
                            break
            if header_row is not None:
                break
    
    # If not found by primary pattern, search for row with most text columns
    if header_row is None:
        best_row = None
        best_score = 0
        
        for idx in range(min(20, len(df))):
            row_data = df.iloc[idx].tolist()
            if is_likely_header_row(row_data, required_headers):
                # Score based on non-empty cells and text content
                non_empty = sum(1 for val in row_data if pd.notna(val) and str(val).strip())
                text_cells = sum(1 for val in row_data 
                               if pd.notna(val) and str(val).strip() and 
                               not str(val).strip().replace('.', '').replace('-', '').isdigit())
                
                score = text_cells * 2 + non_empty  # Prefer text over numbers
                
                if score > best_score:
                    best_score = score
                    best_row = idx
                    # Use first non-empty cell as detected primary
                    for val in row_data:
                        if pd.notna(val) and str(val).strip():
                            detected_primary = str(val).strip()
                            break
        
        header_row = best_row
    
    if header_row is None:
        raise ValueError("Tidak dapat menemukan baris header yang valid dalam 20 baris pertama file Excel")
    
    # Extract headers from the found row
    excel_headers = [normalize_column_name(val) for val in df.iloc[header_row]]
    
    # Remove empty headers from the end
    while excel_headers and not excel_headers[-1]:
        excel_headers.pop()
    
    # Perform strict column matching with enhanced validation
    try:
        valid_headers_mapping, missing_headers, match_details = strict_column_match(excel_headers, required_headers)
        
        logger.info(f"Header row found at: {header_row + 1}")
        logger.info(f"Primary header detected: '{detected_primary}'")
        logger.info(f"Strict validation passed: {len(valid_headers_mapping)}/{len(required_headers)} columns matched")
        
        # Log successful matches
        for excel_header, db_header in valid_headers_mapping:
            if excel_header != db_header:
                logger.info(f"  Mapped: '{excel_header}' -> '{db_header}'")
            else:
                logger.info(f"  Exact match: '{excel_header}'")
        
        return header_row, valid_headers_mapping, missing_headers, detected_primary
        
    except ValueError as e:
        # Re-raise with additional context
        logger.error(f"Column validation failed: {str(e)}")
        raise ValueError(str(e))

def find_data_start_row(df, header_row, detected_primary_header):
    """
    Mencari baris mulai data berdasarkan kolom primary header yang terdeteksi
    Args:
        df: DataFrame Excel
        header_row: Index baris header
        detected_primary_header: Header utama yang terdeteksi
    
    Returns: data_start_row_index
    """
    
    def normalize_text(text):
        if pd.isna(text):
            return ""
        return str(text).lower().strip()

    # Cari index kolom primary header
    excel_headers = []
    for val in df.iloc[header_row]:
        excel_headers.append(str(val).strip() if pd.notna(val) else "")
    
    primary_col_index = None
    
    # Cari kolom yang cocok dengan detected primary header
    for idx, header in enumerate(excel_headers):
        header_norm = normalize_text(header)
        primary_norm = normalize_text(detected_primary_header)
        
        if header_norm and (
            header_norm == primary_norm or
            primary_norm in header_norm or
            header_norm in primary_norm
        ):
            primary_col_index = idx
            break
    
    # Jika tidak ditemukan, gunakan kolom pertama yang tidak kosong
    if primary_col_index is None:
        for idx, header in enumerate(excel_headers):
            if header:
                primary_col_index = idx
                logger.warning(f"Primary header tidak ditemukan, menggunakan kolom pertama: '{header}'")
                break
    
    if primary_col_index is None:
        raise ValueError("Tidak dapat menentukan kolom primary untuk mencari data")
    
    # Cari baris pertama yang memiliki data pada kolom primary
    data_start_row = None
    for idx in range(header_row + 1, len(df)):
        if idx >= len(df):
            break
            
        cell_value = str(df.iloc[idx, primary_col_index]).strip()
        if cell_value and cell_value.lower() not in ['nan', 'none', '', 'null']:
            # Pastikan ini bukan baris kosong atau header tambahan
            row_data = [str(val).strip() for val in df.iloc[idx] if pd.notna(val)]
            if len(row_data) >= 2:  # Minimal 2 kolom berisi data
                data_start_row = idx
                break
    
    if data_start_row is None:
        raise ValueError(f"Tidak ditemukan data setelah header pada kolom '{detected_primary_header}'")
    
    logger.info(f"Data mulai dari baris: {data_start_row + 1} (Excel row: {data_start_row + 2})")
    return data_start_row

def get_column_info(table_name, exclude_automatic=True):
    """
    Mendapatkan informasi kolom dari tabel database dengan informasi lengkap
    Args:
        table_name: Nama tabel
        exclude_automatic: True untuk mengecualikan kolom otomatis (period_date, upload_date)
    """
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = """
        SELECT 
            c.COLUMN_NAME,
            c.DATA_TYPE,
            c.CHARACTER_MAXIMUM_LENGTH,
            c.NUMERIC_PRECISION,
            c.NUMERIC_SCALE,
            c.IS_NULLABLE,
            c.COLUMN_DEFAULT
        FROM INFORMATION_SCHEMA.COLUMNS c
        WHERE c.TABLE_NAME = ? AND c.TABLE_SCHEMA = 'dbo'
        ORDER BY c.ORDINAL_POSITION
        """
        
        cursor.execute(query, (table_name,))
        columns = cursor.fetchall()
        
        if not columns:
            raise ValueError(f"Tabel '{table_name}' tidak ditemukan atau tidak memiliki kolom")
        
        columns_info = {}
        automatic_columns = ['period_date', 'upload_date', 'id']  # Kolom yang otomatis ditambahkan
        
        for col in columns:
            column_name = col[0]
            
            # Skip kolom otomatis jika diminta
            if exclude_automatic and column_name.lower() in automatic_columns:
                continue
                
            columns_info[column_name] = {
                'name': column_name,  # Tambahkan nama kolom untuk error messages
                'data_type': col[1],
                'max_length': col[2],
                'precision': col[3],
                'scale': col[4],
                'is_nullable': col[5] == 'YES',
                'default_value': col[6]
            }
        
        return columns_info
        
    except Exception as e:
        logger.error(f"Error getting column info: {str(e)}")
        raise
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def get_automatic_columns():
    """
    Mendapatkan daftar kolom yang otomatis ditambahkan sistem
    """
    return ['id', 'period_date', 'upload_date']

def validate_and_convert_value(value, column_info, column_name):
    """
    Improved validation with better error handling and type conversion
    """
    try:
        # Handle NULL values first
        processed_value = handle_null_values_for_column(value, {**column_info, 'name': column_name})
        
        # Skip validation for database default marker
        if processed_value == '__USE_DATABASE_DEFAULT__':
            return processed_value, True, ""
        
        # If processed value is None (valid NULL), return it
        if processed_value is None:
            return None, True, ""
        
        # Get column type for conversion
        col_type = column_info.get('data_type', '').upper()
        
        # Convert based on data type with improved error handling
        try:
            if col_type in ['VARCHAR', 'NVARCHAR', 'CHAR', 'NCHAR', 'TEXT']:
                str_value = str(processed_value).strip()
                max_length = column_info.get('max_length')
                
                if max_length and len(str_value) > max_length:
                    return None, False, f"String length ({len(str_value)}) exceeds maximum length ({max_length})"
                
                return str_value, True, ""
                
            elif col_type == 'BIT':
                if isinstance(processed_value, bool):
                    return processed_value, True, ""
                
                str_val = str(processed_value).lower().strip()
                if str_val in ['1', 'true', 'yes', 'y', 'on']:
                    return True, True, ""
                elif str_val in ['0', 'false', 'no', 'n', 'off']:
                    return False, True, ""
                else:
                    return None, False, f"Invalid boolean value: '{processed_value}'. Expected: 1/0, true/false, yes/no"
                    
            elif col_type in ['INT', 'BIGINT', 'SMALLINT', 'TINYINT']:
                # Handle string numbers with better parsing
                if isinstance(processed_value, str):
                    # Remove common thousand separators and currency symbols
                    cleaned = re.sub(r'[,$\s]', '', processed_value.strip())
                    processed_value = cleaned
                
                try:
                    int_value = int(float(processed_value))
                    return int_value, True, ""
                except (ValueError, TypeError, OverflowError):
                    return None, False, f"Invalid integer value: '{processed_value}'"
                    
            elif col_type in ['DECIMAL', 'NUMERIC', 'FLOAT', 'REAL', 'MONEY']:
                if isinstance(processed_value, str):
                    # Remove common thousand separators and currency symbols
                    cleaned = re.sub(r'[,$\s]', '', processed_value.strip())
                    processed_value = cleaned
                
                try:
                    float_value = float(processed_value)
                    # Check for infinity and NaN
                    if not math.isfinite(float_value):
                        return None, False, f"Invalid numeric value (infinity/NaN): '{processed_value}'"
                    return float_value, True, ""
                except (ValueError, TypeError):
                    return None, False, f"Invalid numeric value: '{processed_value}'"
                    
            elif col_type in ['DATE', 'DATETIME', 'DATETIME2', 'SMALLDATETIME']:
                if isinstance(processed_value, (datetime, date)):
                    return processed_value, True, ""
                
                # Try to parse string dates with more formats
                date_formats = [
                    '%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f',
                    '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d', 
                    '%d-%m-%Y', '%m-%d-%Y', '%Y%m%d',
                    '%d.%m.%Y', '%m.%d.%Y',
                    '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M:%S',
                    '%m/%d/%Y %H:%M:%S', '%Y/%m/%d %H:%M:%S'
                ]
                
                processed_str = str(processed_value).strip()
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(processed_str, fmt)
                        return parsed_date.date() if col_type == 'DATE' else parsed_date, True, ""
                    except ValueError:
                        continue
                
                return None, False, f"Invalid date format: '{processed_value}'. Expected formats: YYYY-MM-DD, MM/DD/YYYY, etc."
                
            else:
                # For unknown types, return as string with validation
                str_value = str(processed_value).strip()
                return str_value, True, ""
                
        except Exception as convert_error:
            return None, False, f"Type conversion error: {str(convert_error)}"
            
    except Exception as e:
        return None, False, f"Validation error: {str(e)}"

def validate_batch_data(df, columns_info):
    """
    Validate entire batch of data before insert
    Returns: (validated_df, is_valid, validation_errors)
    PERBAIKAN: Handle database default marker
    """
    validation_errors = []
    validated_data = {}
    
    # Initialize validated data structure
    for col in df.columns:
        validated_data[col] = []
    
    # Validate each row
    for idx, row in df.iterrows():
        row_errors = []
        row_data = {}
        
        # Validate each column in the row
        for col in df.columns:
            if col in columns_info:
                value = row[col]
                column_info = columns_info[col]
                
                converted_value, is_valid, error_msg = validate_and_convert_value(
                    value, column_info, col
                )
                
                if not is_valid:
                    row_errors.append(f"Row {idx + 1}, Column '{col}': {error_msg}")
                else:
                    row_data[col] = converted_value
            else:
                # PERBAIKAN: Handle kolom yang tidak ada di schema database
                # Ini mungkin kolom yang akan di-skip atau kolom otomatis
                row_data[col] = row[col]
        
        # If any validation error in this row, record it
        if row_errors:
            validation_errors.extend(row_errors)
        
        # PERBAIKAN: Selalu tambahkan data row, bahkan jika ada error
        # Ini untuk mempertahankan struktur DataFrame
        for col in df.columns:
            if col in row_data:
                validated_data[col].append(row_data[col])
            else:
                validated_data[col].append(row[col])  # Use original value if not processed
    
    # If any validation errors, return failure
    if validation_errors:
        return None, False, validation_errors
    
    # Create validated DataFrame
    validated_df = pd.DataFrame(validated_data)
    return validated_df, True, []

def process_uploaded_data(df, table_name):
    """
    Process uploaded DataFrame with proper NULL handling
    Improved version with better error reporting
    """
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get column information for the table
        cursor.execute("""
            SELECT 
                c.COLUMN_NAME,
                c.DATA_TYPE,
                c.IS_NULLABLE,
                c.COLUMN_DEFAULT,
                c.CHARACTER_MAXIMUM_LENGTH
            FROM INFORMATION_SCHEMA.COLUMNS c
            WHERE c.TABLE_NAME = ?
            AND c.COLUMN_NAME NOT IN ('id', 'period_date', 'upload_date')
            ORDER BY c.ORDINAL_POSITION
        """, (table_name,))
        
        columns_info = cursor.fetchall()
        
        # Create column mapping
        column_mapping = {}
        for col_info in columns_info:
            column_mapping[col_info[0]] = {
                'name': col_info[0],
                'data_type': col_info[1],
                'is_nullable': col_info[2] == 'YES',
                'default_value': col_info[3],
                'max_length': col_info[4]
            }
        
        # Process each row in DataFrame
        processed_rows = []
        for index, row in df.iterrows():
            processed_row = {}
            
            for col_name, value in row.items():
                if col_name in column_mapping:
                    try:
                        processed_value = handle_null_values_for_column(
                            value, 
                            column_mapping[col_name]
                        )
                        processed_row[col_name] = processed_value
                    except ValueError as ve:
                        logger.error(f"Row {index + 1}, Column {col_name}: {str(ve)}")
                        raise ValueError(f"Row {index + 1}, Column {col_name}: {str(ve)}")
            
            processed_rows.append(processed_row)
        
        return processed_rows
        
    except Exception as e:
        logger.error(f"Error processing uploaded data: {str(e)}")
        raise
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def process_excel_file(file_path, table_name, primary_header=None, sheet_name=None, periode_date=None):
    """
    Enhanced Excel processing with strict column validation
    """
    try:
        if not sheet_name:
            return {'success': False, 'message': 'Nama sheet harus dipilih'}

        # Read Excel file
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        except ValueError as e:
            if "Worksheet named" in str(e):
                return {'success': False, 'message': f'Sheet "{sheet_name}" tidak ditemukan dalam file Excel'}
            else:
                return {'success': False, 'message': f'Error membaca sheet: {str(e)}'}

        if df.empty:
            return {'success': False, 'message': f'Sheet "{sheet_name}" kosong atau tidak memiliki data'}

        # Get database column information
        columns_info = get_column_info(table_name, exclude_automatic=True)
        required_headers = list(columns_info.keys())

        # Enhanced header validation with strict checking
        try:
            header_row, valid_headers_mapping, missing_headers, detected_primary = find_header_row_and_validate(
                df, required_headers, primary_header
            )
        except ValueError as validation_error:
            # Extract actual Excel headers for error display
            excel_headers_for_error = []
            if len(df) > 0:
                # Try to find the most likely header row for error display
                for idx in range(min(10, len(df))):
                    row_data = [str(val).strip() if pd.notna(val) else "" for val in df.iloc[idx]]
                    non_empty_headers = [h for h in row_data if h]
                    if len(non_empty_headers) >= 2:  # At least 2 non-empty headers
                        excel_headers_for_error = row_data
                        break
                
                # If no good header row found, use first row
                if not excel_headers_for_error:
                    excel_headers_for_error = [str(val).strip() if pd.notna(val) else "" for val in df.iloc[0]]
            
            return {
                'success': False,
                'message': str(validation_error),
                'validation_type': 'column_structure',
                'header_info': {
                    'required_headers': required_headers,
                    'excel_headers': excel_headers_for_error
                }
            }
        
        logger.info(f"Strict validation passed: {len(valid_headers_mapping)} columns matched")
        
        # Continue with rest of the processing...
        # Find data start row
        data_start_row = find_data_start_row(df, header_row, detected_primary)

        # Extract and map data
        excel_headers = [normalize_column_name(val) for val in df.iloc[header_row]]
        data_df = df.iloc[data_start_row:].copy()
        data_df.columns = range(len(data_df.columns))

        # Create column index mapping
        col_index_mapping = {}
        for excel_header, db_header in valid_headers_mapping:
            for idx, header in enumerate(excel_headers):
                if header == excel_header:
                    col_index_mapping[idx] = db_header
                    break

        # Extract mapped data
        filtered_data = {}
        for col_idx, db_header in col_index_mapping.items():
            if col_idx < len(data_df.columns):
                filtered_data[db_header] = data_df.iloc[:, col_idx]

        # Add default values for missing columns (should be none in strict mode)
        for missing_col in missing_headers:
            if missing_col in columns_info:
                col_info = columns_info[missing_col]
                if col_info.get('default_value') is not None:
                    default_val = process_default_value(col_info['default_value'], col_info)
                    filtered_data[missing_col] = [default_val] * len(data_df)
                elif col_info.get('is_nullable', False):
                    filtered_data[missing_col] = [None] * len(data_df)
                else:
                    logger.warning(f"Missing required column '{missing_col}' has no default value and is not nullable")

        if not filtered_data:
            raise ValueError("Tidak ada data yang dapat diekstrak dari file Excel")

        # Create final DataFrame
        final_df = pd.DataFrame(filtered_data)
        
        # Clean empty rows
        final_df = final_df.dropna(how='all')
        final_df = final_df[~final_df.astype(str).apply(lambda x: x.str.strip().eq('').all(), axis=1)]
        final_df = final_df.reset_index(drop=True)

        if len(final_df) == 0:
            return {
                'success': False,
                'message': 'Tidak ada data valid ditemukan untuk diinsert',
                'header_info': {
                    'header_row': header_row + 1,
                    'data_start_row': data_start_row + 1,
                    'detected_primary': detected_primary,
                    'found_headers': [db_header for _, db_header in valid_headers_mapping],
                    'missing_headers': missing_headers,
                    'sheet_used': sheet_name
                }
            }

        # Continue with validation and database insertion...
        relevant_columns_info = {col: columns_info[col] for col in final_df.columns if col in columns_info}
        
        validation_errors = []
        validated_data = {}
        
        # Initialize validated data structure
        for col in final_df.columns:
            validated_data[col] = []
        
        # Validate each row with improved error handling
        for idx, row in final_df.iterrows():
            row_data = {}
            
            for col in final_df.columns:
                if col in relevant_columns_info:
                    value = row[col]
                    column_info = relevant_columns_info[col]
                    
                    converted_value, is_valid, error_msg = validate_and_convert_value(
                        value, column_info, col
                    )
                    
                    if not is_valid:
                        validation_errors.append(f"Row {idx + 1}, Column '{col}': {error_msg}")
                        row_data[col] = None
                    else:
                        row_data[col] = converted_value
                else:
                    row_data[col] = row[col]
            
            # Add row data
            for col in final_df.columns:
                validated_data[col].append(row_data.get(col, row[col]))
        
        # Create validated DataFrame
        validated_df = pd.DataFrame(validated_data)
        
        # Enhanced validation error handling
        if validation_errors:
            error_rate = len(validation_errors) / (len(final_df) * len(final_df.columns))
            if error_rate > 0.1:  # More than 10% error rate
                return {
                    'success': False,
                    'message': f'Validasi data gagal. Tingkat kesalahan: {error_rate:.1%}',
                    'validation_errors': validation_errors[:20],
                    'total_errors': len(validation_errors),
                    'rows_processed': len(final_df),
                    'validation_type': 'data_validation'
                }

        # Process uploaded data
        processed_data = process_uploaded_data(validated_df, table_name)
        processed_df = pd.DataFrame(processed_data)

        # Insert to database
        result = insert_to_database(processed_df, table_name, periode_date, replace_existing=True)

        # Add detailed results
        result['header_info'] = {
            'periode_date': periode_date,
            'header_row': header_row + 1,
            'data_start_row': data_start_row + 1,
            'detected_primary': detected_primary,
            'found_headers': [db_header for _, db_header in valid_headers_mapping],
            'missing_headers': missing_headers,
            'sheet_used': sheet_name,
            'matched_columns': len(valid_headers_mapping),
            'required_columns': len(required_headers),
            'validation_warnings': len(validation_errors)
        }

        if validation_errors and len(validation_errors) <= 10:
            result['validation_warnings'] = validation_errors

        return result

    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        return {'success': False, 'message': f'Error processing file: {str(e)}'}

def insert_to_database(df, table_name, periode_date=None, replace_existing=True):
    """
    Insert dataframe ke SQL Server table - data sudah tervalidasi
    Includes period_date and upload_date automatic columns
    PERBAIKAN: Handle database default values
    """
    conn = None
    cursor = None

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # PERBAIKAN: Filter kolom yang menggunakan database default
        # Check which columns have database defaults
        columns_with_defaults = {}
        for col in df.columns:
            # Check if any row in this column has the database default marker
            has_default_marker = df[col].astype(str).str.contains('__USE_DATABASE_DEFAULT__').any()
            if has_default_marker:
                columns_with_defaults[col] = True

        logger.info(f"Kolom dengan database default: {list(columns_with_defaults.keys())}")

        # PERBAIKAN: Build dynamic column list per row
        successful_inserts = 0
        current_datetime = datetime.now()
        
        # Delete existing data if replace_existing is True
        if replace_existing and periode_date:
            delete_query = f"DELETE FROM {table_name} WHERE period_date = ?"
            cursor.execute(delete_query, (periode_date,))
            logger.info(f"Data sebelumnya dengan periode {periode_date} telah dihapus dari {table_name}")

        # PERBAIKAN: Process each row individually to handle different column sets
        for idx, row in df.iterrows():
            try:
                # Build column list and values for this specific row
                insert_columns = []
                insert_values = []
                placeholders = []

                # Process data columns
                for col in df.columns:
                    raw_value = row[col]
                    
                    # PERBAIKAN: Skip kolom jika menggunakan database default
                    if str(raw_value) == '__USE_DATABASE_DEFAULT__':
                        logger.debug(f"Row {idx + 1}: Skipping column '{col}' - using database default")
                        continue
                    
                    # Include column in insert
                    insert_columns.append(f'[{col}]')
                    placeholders.append('?')
                    
                    # Normalize value
                    dtype = str(df[col].dtype)
                    value = normalize_value(raw_value, dtype)
                    insert_values.append(value)

                # Add automatic columns
                insert_columns.extend(['[period_date]', '[upload_date]'])
                placeholders.extend(['?', 'GETDATE()'])
                insert_values.append(periode_date)  # period_date value
                # upload_date uses GETDATE() in SQL, no value needed

                if not insert_columns:
                    logger.warning(f"Row {idx + 1}: No columns to insert, skipping")
                    continue

                # Build and execute query for this row
                insert_query = f"INSERT INTO {table_name} ({', '.join(insert_columns)}) VALUES ({', '.join(placeholders)})"
                
                logger.debug(f"Row {idx + 1} Query: {insert_query}")
                logger.debug(f"Row {idx + 1} Values: {insert_values}")
                
                cursor.execute(insert_query, insert_values)
                successful_inserts += 1

            except Exception as row_error:
                logger.error(f"Error inserting row {idx + 1}: {str(row_error)}")
                # Continue with next row instead of failing completely
                continue

        conn.commit()
        logger.info(f"Berhasil insert {successful_inserts} dari {len(df)} baris")

        return {
            'success': True,
            'message': f'Berhasil insert {successful_inserts} baris data',
            'inserted_rows': successful_inserts,
            'skipped_rows': len(df) - successful_inserts,
            'error_rows': 0,
            'columns_used': [col for col in df.columns.tolist() if col not in columns_with_defaults],
            'columns_with_defaults': list(columns_with_defaults.keys()),
            'periode_date': periode_date,
            'upload_date': current_datetime.strftime('%Y-%m-%d %H:%M:%S')
        }

    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error inserting to database: {str(e)}")
        return {
            'success': False,
            'message': f'Error saat insert ke database: {str(e)}',
            'inserted_rows': 0,
            'skipped_rows': 0,
            'error_rows': len(df) if 'df' in locals() else 0
        }
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def normalize_value(value, dtype=None):
    """
    Normalisasi nilai untuk insert ke SQL Server:
    - 'N/A', '', 'None', 'null', NaN => None (untuk kolom non-numeric)
    - '-' => 0 jika numeric, else None
    - Membersihkan tanda kurung & kutip jika ada
    PERBAIKAN: Handle database default marker
    """
    # PERBAIKAN: Handle database default marker - return as-is tanpa processing
    if str(value) == '__USE_DATABASE_DEFAULT__':
        return value
    
    if pd.isna(value):
        return 0 if dtype and ('int' in dtype or 'float' in dtype) else None

    cleaned = str(value).strip()
    cleaned_simple = cleaned.strip("'\"() ").upper()

    null_equivalents = {'N/A', '', 'NONE', 'NULL'}

    if cleaned_simple in null_equivalents:
        return 0 if dtype and ('int' in dtype or 'float' in dtype) else None

    if cleaned == '-':
        return 0 if dtype and ('int' in dtype or 'float' in dtype) else None

    return value
    
def process_default_value(default_value, column_info):
    """
    Process default value based on column type
    PERBAIKAN: Fungsi ini sekarang hanya untuk generate nilai default eksplisit
    """
    col_type = column_info.get('data_type', '').upper()
    
    # PERBAIKAN: Jangan proses default value yang berbentuk SQL function atau constraint
    default_str = str(default_value).strip()
    
    # Skip SQL Server constraint format seperti ((1)), (getdate()), etc.
    if default_str.startswith('((') and default_str.endswith('))'):
        return '__USE_DATABASE_DEFAULT__'
    elif default_str.startswith('(') and default_str.endswith(')'):
        # Check if it's a SQL function
        inner_value = default_str[1:-1].strip()
        if inner_value.upper() in ['GETDATE', 'GETUTCDATE', 'SYSDATETIME', 'CURRENT_TIMESTAMP']:
            return '__USE_DATABASE_DEFAULT__'
        # If it's just wrapped value like (1), extract it
        default_str = inner_value
    
    try:
        if col_type in ['VARCHAR', 'NVARCHAR', 'CHAR', 'NCHAR', 'TEXT']:
            # Remove quotes if present
            if default_str.startswith("'") and default_str.endswith("'"):
                default_str = default_str[1:-1]
            return str(default_str)
            
        elif col_type == 'BIT':
            if default_str.lower() in ['1', 'true']:
                return True
            elif default_str.lower() in ['0', 'false']:
                return False
            else:
                return bool(int(default_str))
                
        elif col_type in ['INT', 'BIGINT', 'SMALLINT', 'TINYINT']:
            return int(float(default_str))
            
        elif col_type in ['DECIMAL', 'NUMERIC', 'FLOAT', 'REAL']:
            return float(default_str)
            
        elif col_type in ['DATE', 'DATETIME', 'DATETIME2']:
            # PERBAIKAN: Untuk date/datetime, selalu gunakan database default jika ada function
            if default_str.upper() in ['GETDATE()', 'GETUTCDATE()', 'SYSDATETIME()', 'CURRENT_TIMESTAMP']:
                return '__USE_DATABASE_DEFAULT__'
            else:
                # Try to parse date string
                from datetime import datetime
                date_formats = ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%m/%d/%Y', '%d/%m/%Y']
                for fmt in date_formats:
                    try:
                        return datetime.strptime(default_str, fmt)
                    except ValueError:
                        continue
                # If can't parse, let database handle it
                return '__USE_DATABASE_DEFAULT__'
                
        elif col_type == 'TIME':
            from datetime import datetime, time
            time_formats = ['%H:%M:%S', '%H:%M']
            for fmt in time_formats:
                try:
                    parsed_time = datetime.strptime(default_str, fmt).time()
                    return parsed_time
                except ValueError:
                    continue
            return '__USE_DATABASE_DEFAULT__'
            
        else:
            return str(default_str)
            
    except (ValueError, TypeError) as e:
        logger.warning(f"Error processing default value '{default_value}' for column '{column_info['name']}': {str(e)}")
        return '__USE_DATABASE_DEFAULT__'

def handle_null_values_for_column(value, column_info):
    """
    Handle NULL values based on column configuration
    Improved version with better type handling and default value processing
    """
    # Check if value is considered "null" in various formats
    null_indicators = [None, '', 'NULL', 'null', 'Null', 'N/A', 'n/a', 'NA', 'na', '#N/A']
    
    # Handle pandas NaN and empty strings more thoroughly
    is_null = (value in null_indicators or 
               (isinstance(value, float) and pd.isna(value)) or
               (isinstance(value, str) and value.strip() in [''] + null_indicators) or
               value is None)
    
    if is_null:
        if column_info.get('is_nullable', False):
            return None  # Return None untuk NULL database value
        else:
            default_value = column_info.get('default_value')
            if default_value is not None and str(default_value).strip() != '':
                # INI SATU-SATUNYA TEMPAT PEMANGGILAN process_default_value
                return process_default_value(default_value, column_info)
            else:
                col_name = column_info.get('name', 'Unknown')
                raise ValueError(f"Column '{col_name}' cannot be NULL and has no default value")
    
    # If not null, return the value as is (will be processed later based on column type)
    return value

def get_template_tables(role_access=None, division=None):
    """
    Mendapatkan daftar template dari MasterCreator.
    Jika role_access = 'user', hanya kembalikan template yang sesuai division user.
    """
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if role_access and role_access.lower() == 'user':
            cursor.execute("""
                SELECT template_name
                FROM MasterCreator
                WHERE division_name = ?
                ORDER BY create_date DESC
            """, (division,))
        else:
            cursor.execute("""
                SELECT template_name
                FROM MasterCreator
                ORDER BY create_date DESC
            """)
        
        tables = cursor.fetchall()
        return [table[0] for table in tables]
        
    except Exception as e:
        logger.error(f"Error getting template tables: {str(e)}")
        return []
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def get_data_count_by_period(table_name):
    """
    Mendapatkan jumlah data berdasarkan period_date untuk tabel tertentu
    """
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if table exists and has period_date column
        check_query = """
        SELECT COUNT(*) 
        FROM INFORMATION_SCHEMA.COLUMNS 
        WHERE TABLE_NAME = ? AND COLUMN_NAME = 'period_date' AND TABLE_SCHEMA = 'dbo'
        """
        cursor.execute(check_query, (table_name,))
        has_period_date = cursor.fetchone()[0] > 0
        
        if not has_period_date:
            return []
        
        # Get data count grouped by period_date
        query = f"""
        SELECT TOP 5
            period_date,
            COUNT(*) as total_records
        FROM [{table_name}]
        GROUP BY period_date
        ORDER BY period_date DESC
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        
        data_counts = []
        for row in results:
            data_counts.append({
                'period_date': row[0].strftime('%B %Y') if row[0] else 'NULL',
                'total_records': row[1]
            })
        
        return data_counts
        
    except Exception as e:
        logger.error(f"Error getting data count by period: {str(e)}")
        return []
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
            
def get_master_divisions_tables():
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT division_name 
            FROM MasterDivisions
        """)
        
        tables = cursor.fetchall()
        return [table[0] for table in tables]
        
    except Exception as e:
        logger.error(f"Error getting template tables: {str(e)}")
        return []
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def get_excel_sheets(file_path):
    """
    Mendapatkan daftar nama sheet dalam file Excel
    Args:
        file_path: Path ke file Excel
    Returns:
        List nama sheet atau None jika error
    """
    try:
        # Baca file Excel untuk mendapatkan daftar sheet
        excel_file = pd.ExcelFile(file_path)
        sheets = excel_file.sheet_names
        excel_file.close()
        
        logger.info(f"Ditemukan {len(sheets)} sheet: {sheets}")
        return sheets
        
    except Exception as e:
        logger.error(f"Error reading Excel sheets: {str(e)}")
        return None

def convert_value_for_sql_server(value):
    """
    Convert Python values specifically for SQL Server driver compatibility
    """
    if value is None:
        return None
    
    # Boolean conversion - critical for SQL Server driver
    elif isinstance(value, bool):
        return 1 if value else 0
    
    # String handling - ensure proper encoding
    elif isinstance(value, str):
        # Remove null bytes yang bisa menyebabkan error
        return value.replace('\x00', '')
    
    # Numeric types
    elif isinstance(value, (int, float)):
        return value
    
    # Decimal handling
    elif isinstance(value, Decimal):
        return float(value)
    
    # DateTime handling
    elif isinstance(value, (datetime, date)):
        return value
    
    # Complex types - convert to string
    elif isinstance(value, (list, dict, tuple)):
        return str(value)
    
    # Bytes handling
    elif isinstance(value, bytes):
        return value
    
    # Default case
    else:
        return str(value)

def safe_insert_single_record(table_name, columns, values):
    """
    Helper function untuk insert single record dengan error handling
    """
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Convert values
        converted_values = [convert_value_for_sql_server(value) for value in values]
        
        # Build query
        placeholders = ', '.join(['?' for _ in columns])
        column_names = ', '.join(columns)
        query = f"INSERT INTO {table_name} ({column_names}) VALUES ({placeholders})"
        
        # Debug info
        print(f"Inserting to {table_name}:")
        for col, val in zip(columns, converted_values):
            print(f"  {col}: {type(val).__name__} = {repr(val)}")
        
        cursor.execute(query, tuple(converted_values))
        conn.commit()
        print(f"✅ Successfully inserted to {table_name}")
        return True
        
    except Exception as e:
        print(f"❌ Error inserting to {table_name}: {e}")
        if conn:
            conn.rollback()
        return False
        
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def validate_password_strength(password):
    """
    Validasi password dengan kriteria:
    - Minimal 9 karakter
    - Mengandung huruf kapital
    - Mengandung angka
    - Mengandung karakter spesial
    """
    if len(password) < 9:
        return False, "Password minimal 9 karakter"
    
    if not re.search(r'[A-Z]', password):
        return False, "Password harus mengandung huruf kapital"
    
    if not re.search(r'\d', password):
        return False, "Password harus mengandung angka"
    
    if not re.search(r'[@$!%*?&]', password):
        return False, "Password harus mengandung karakter spesial (@$!%*?&)"
    
    return True, "Password valid"

         
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("SELECT password_hash, role_access, division, fullname FROM MasterUsers WHERE username = ?", (username,))
        user = cur.fetchone()

        if user and bcrypt.check_password_hash(user[0], password):
            session['username'] = username
            session['division'] = user[2]
            session['fullname'] = user[3]
            session['role_access'] = user[1]
            session['upload_done'] = True
            return redirect(url_for('upload_file'))
        
        flash("Invalid username or password.")

    return render_template('login.html')

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    conn = get_db_connection()
    cur = conn.cursor()
        
    if 'username' not in session:
        flash("You need to log in first.")
        return redirect(url_for('login'))
    
    role_access = session.get('role_access')
    fullname = session.get('fullname')
    username = session.get('username')
    division = session.get('division')
    
    if request.method == 'GET':
        return render_template(
            'change_password.html',
            username=username,
            division=division,
            role_access=role_access,
            fullname=fullname
        )

    elif request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        password_confirm = request.form['password_confirm']
        username = session['username']

        # Validasi password confirmation
        if new_password != password_confirm:
            return render_template(
                'change_password.html',
                username=username,
                division=division,
                role_access=role_access,
                fullname=fullname,
                error_confirm="Passwords do not match.",
                current_password=current_password,
                new_password=new_password,
                password_confirm=password_confirm
            )

        # Validasi password strength
        is_valid, error_message = validate_password_strength(new_password)
        if not is_valid:
            return render_template(
                'change_password.html',
                username=username,
                division=division,
                role_access=role_access,
                fullname=fullname,
                error_strength=error_message,
                current_password=current_password,
                new_password=new_password,
                password_confirm=password_confirm
            )

        # Fetch the current hashed password from the database
        cur.execute("SELECT password_hash FROM MasterUsers WHERE username = ?", (username,))
        user = cur.fetchone()

        if not user or not bcrypt.check_password_hash(user[0], current_password):
            return render_template(
                'change_password.html',
                username=username,
                division=division,
                role_access=role_access,
                fullname=fullname,
                error_current="Current password is incorrect.",
                current_password=current_password,
                new_password=new_password,
                password_confirm=password_confirm
            )
            
        else:
            # Hash the new password and update the database
            new_password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
            cur.execute("UPDATE MasterUsers SET password_hash = ? WHERE username = ?", (new_password_hash, username))
            conn.commit()
            
            return '''
                <script>
                    alert("Perubahan password berhasil dilakukan.");
                    window.location.href = "{}";
                </script>
            '''.format(url_for('upload_file'))

# Logout Route
@app.route('/logout')
def logout():
    session.pop('username', None)
    flash("You have been logged out.")
    return redirect(url_for('login'))

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if 'username' not in session:
        flash("Please log in first.")
        return redirect(url_for('login'))

    role_access = session.get('role_access')
    fullname = session.get('fullname')
    username = session.get('username')
    division = session.get('division')
    
    if request.method == 'GET':
        # Mendapatkan daftar tabel template untuk dropdown
        template_tables = get_template_tables(role_access, division)
        
        if role_access.lower() == 'user':
            # Ambil hanya tabel yang sesuai division
            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT table_name 
                FROM INFORMATION_SCHEMA.TABLES t
                JOIN MasterDivisions d ON t.TABLE_SCHEMA = 'dbo'
                WHERE d.division_name = ?
            """, (division,))
            allowed_tables = [row[0] for row in cursor.fetchall()]
            cursor.close()
            conn.close()

            # Hanya ambil table yang masuk allowed_tables
            template_tables = [tbl for tbl in template_tables if tbl in allowed_tables]
        
        # Tampilkan halaman upload dengan data tabel template
        return render_template(
            'upload.html',
            username=username,
            division=division,
            role_access=role_access,
            fullname=fullname,
            template_tables=template_tables
        )
    
    elif request.method == 'POST':
        try:
            if 'file' not in request.files:
                return jsonify({'success': False, 'message': 'Tidak ada file yang dipilih'})
            
            file = request.files['file']
            table_name = request.form.get('table_name', '').strip()
            primary_header = request.form.get('primary_header', '').strip() or None
            sheet_name = request.form.get('sheet_name', '').strip() or None
            periode_date = request.form.get('periode_date', '').strip() or None
            
            if file.filename == '':
                return jsonify({'success': False, 'message': 'Tidak ada file yang dipilih'})
            
            if not table_name:
                return jsonify({'success': False, 'message': 'Nama tabel harus dipilih'})
            
            if not allowed_file(file.filename):
                return jsonify({'success': False, 'message': 'File harus berformat Excel (.xlsx atau .xls)'})
            
            # Validasi format tanggal periode
            if periode_date:
                try:
                    periode_date = datetime.strptime(periode_date, '%Y-%m').date().replace(day=1)
                except ValueError:
                    return jsonify({'success': False, 'message': 'Format tanggal periode tidak valid. Gunakan format YYYY-MM'})
            
            # Buat folder uploads ada
            upload_folder = app.config['UPLOAD_FOLDER']
            os.makedirs(upload_folder, exist_ok=True)
            
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            file_path = os.path.join(upload_folder, filename)
            file.save(file_path)
            
            print(f"Periode Date: {periode_date}")
            # Proses file
            result = process_excel_file(file_path, table_name, primary_header, sheet_name, periode_date)
            
            # Simpan ke tabel MasterUploader dengan error handling yang lebih baik
            try:
                print("Attempting to insert to MasterUploader...")
                
                # Gunakan fungsi safe_insert_single_record
                columns = ['username', 'division', 'template', 'sheets', 'file_upload', 'period_date', 'upload_date']
                values = [
                    session.get('username'),
                    session.get('division'),
                    table_name,
                    sheet_name,
                    file_path,
                    periode_date,
                    datetime.now()
                ]
                
                # Debug data sebelum insert
                print("Data to be inserted to MasterUploader:")
                for col, val in zip(columns, values):
                    print(f"  {col}: {type(val).__name__} = {repr(val)}")
                
                insert_success = safe_insert_single_record('MasterUploader', columns, values)
                
                if not insert_success:
                    logger.warning("Failed to insert to MasterUploader, but continuing with main process")
                    
            except Exception as e:
                logger.error(f"Gagal insert ke MasterUploader: {str(e)}")
                # Don't fail the entire process if MasterUploader insert fails
                logger.warning("Continuing with main process despite MasterUploader insert failure")
            
            return jsonify(result)
        
        except Exception as e:
            logger.error(f"Error in upload_file: {str(e)}")
            return jsonify({'success': False, 'message': f'Error: {str(e)}'})
    
    return redirect(url_for('upload_file'))                

@app.route('/preview-headers/<table_name>')
def preview_headers(table_name):
    """
    Preview header database dengan menandai kolom otomatis
    """
    try:
        # Dapatkan semua kolom termasuk otomatis
        all_columns = get_column_info(table_name, exclude_automatic=False)
        
        # Dapatkan kolom yang diharapkan dari Excel (tanpa otomatis)
        excel_columns = get_column_info(table_name, exclude_automatic=True)
        
        # Tandai kolom otomatis
        automatic_columns = get_automatic_columns()
        
        headers_with_info = {}
        for col_name, col_info in all_columns.items():
            headers_with_info[col_name] = {
                **col_info,
                'is_automatic': col_name.lower() in [ac.lower() for ac in automatic_columns],
                'required_in_excel': col_name in excel_columns
            }
        
        data_counts = get_data_count_by_period(table_name)
        
        return jsonify({
            'success': True,
            'headers': headers_with_info,
            'table_name': table_name,
            'total_columns': len(all_columns),
            'excel_required_columns': len(excel_columns),
            'automatic_columns': len(automatic_columns),
            'data_counts_by_period': data_counts
        })
        
    except Exception as e:
        logger.error(f"Error getting table headers: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })

@app.route('/analyze-excel', methods=['POST'])
def analyze_excel():
    """Analyze Excel file structure without inserting to database"""
    
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first.'})
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Tidak ada file yang dipilih'})
        
        file = request.files['file']
        sheet_name = request.form.get('sheet_name', '').strip() or None
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Tidak ada file yang dipilih'})
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': 'File harus berformat Excel (.xlsx atau .xls)'})
        
        # Simpan file temporary
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"analyze_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        try:
            # Dapatkan daftar sheet terlebih dahulu
            available_sheets = get_excel_sheets(file_path)
            
            # Baca file Excel dengan sheet yang spesifik atau default
            if sheet_name:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    sheet_used = sheet_name
                except ValueError:
                    return jsonify({'success': False, 'message': f'Sheet "{sheet_name}" tidak ditemukan dalam file Excel'})
            else:
                df = pd.read_excel(file_path, header=None)
                sheet_used = available_sheets[0] if available_sheets else 'Sheet pertama'
            
            # Auto-detect header
            header_row, detected_primary = find_primary_header_row(df)
            
            # Get headers
            excel_headers = []
            for val in df.iloc[header_row]:
                if pd.notna(val):
                    excel_headers.append(str(val).strip())
                else:
                    excel_headers.append("")
            
            # Find data start row
            data_start_row = find_data_start_row(df, header_row, detected_primary)
            total_data_rows = len(df) - data_start_row
            
            # Sample data
            sample_data = []
            for i in range(data_start_row, min(data_start_row + 3, len(df))):
                row_sample = []
                for j in range(min(len(excel_headers), len(df.columns))):
                    val = df.iloc[i, j]
                    row_sample.append(str(val) if pd.notna(val) else "")
                sample_data.append(row_sample)
            
            result = {
                'success': True,
                'analysis': {
                    'total_rows': total_data_rows,
                    'total_columns': len(df.columns),
                    'header_row': header_row + 1,
                    'data_start_row': data_start_row + 1,
                    'detected_primary_header': detected_primary,
                    'headers': excel_headers,
                    'sample_data': sample_data,
                    'data_rows_available': len(df) - data_start_row,
                    'available_sheets': available_sheets,
                    'sheet_used': sheet_used
                }
            }
            
        finally:
            # Hapus file temporary
            try:
                os.remove(file_path)
            except:
                pass
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error in analyze_excel: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/get-template-tables')
def get_template_tables_endpoint():
    """Endpoint untuk mendapatkan daftar tabel template"""
    try:
        tables = get_template_tables()
        return jsonify({
            'success': True, 
            'tables': tables,
            'message': f'Ditemukan {len(tables)} tabel template'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# CRUD Template Tables
@app.route('/create-table', methods=['GET', 'POST'])
def create_table():
    if 'username' not in session:
        flash("Please log in first.")
        return redirect(url_for('login'))

    role_access = session.get('role_access')
    fullname = session.get('fullname')
    username = session.get('username')
    division = session.get('division')
    divisions = get_master_divisions_tables()
    
    if request.method == 'GET':
        return render_template(
            'create_table.html',
            username=username,
            division=division,
            role_access=role_access,
            fullname=fullname,
            divisions=divisions
        )
    
    elif request.method == 'POST':
        conn = None
        cursor = None
        
        try:
            data = request.get_json()
            table_name = data.get('table_name', '').strip()
            columns = data.get('columns', [])
            divisions = data.get('divisions', '').strip()
            
            if not table_name:
                return jsonify({'success': False, 'message': 'Nama tabel harus diisi'})
            
            if not columns:
                return jsonify({'success': False, 'message': 'Minimal harus ada satu kolom'})
            
            if not divisions:
                return jsonify({'success': False, 'message': 'Divisi harus dipilih'})
            
            # Validate table name (alphanumeric and underscore only, no spaces)
            if not re.match(r'^[a-zA-Z][a-zA-Z0-9_]*$', table_name):
                return jsonify({'success': False, 'message': 'Nama tabel hanya boleh mengandung huruf, angka, dan underscore. Harus dimulai dengan huruf dan tidak boleh ada spasi.'})
            
            # Get database connection
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Matikan autocommit dan gunakan explicit transaction
            conn.autocommit = False
            
            # Check table existence dengan lock
            cursor.execute("""
                SELECT COUNT(*) FROM sys.tables 
                WHERE name = ? AND type = 'U'
            """, (table_name,))
            
            if cursor.fetchone()[0] > 0:
                return jsonify({'success': False, 'message': f'Tabel "{table_name}" sudah ada dalam database'})
            
            # Build CREATE TABLE query with automatic columns
            create_query = f"CREATE TABLE [{table_name}] (\n"
            create_query += "    [id] INT IDENTITY(1,1) PRIMARY KEY,\n"
            
            column_definitions = []
            for col in columns:
                col_name = col.get('name', '').strip()
                reserved_columns = ['id', 'period_date', 'upload_date']
                
                if col_name.lower() in reserved_columns:
                    return jsonify({'success': False, 'message': f'Kolom "{col_name}" tidak boleh dibuat secara manual karena sudah ditambahkan otomatis'})
                
                col_type = col.get('type', 'VARCHAR')
                col_length = col.get('length', '')
                allow_nulls = col.get('allow_nulls', False)
                default_value = col.get('default_value', None)
                
                if not col_name:
                    return jsonify({'success': False, 'message': 'Semua kolom harus memiliki nama'})
                
                # Validate column name - DIPERBAIKI untuk menolak spasi
                if not re.match(r'^[a-zA-Z][a-zA-Z0-9_]*$', col_name):
                    return jsonify({'success': False, 'message': f'Nama kolom "{col_name}" tidak valid. Hanya boleh huruf, angka, dan underscore (tanpa spasi).'})
                
                # Build column definition
                col_def = f"    [{col_name}] {col_type}"
                
                # Add length for applicable types
                if col_type in ['VARCHAR', 'NVARCHAR', 'CHAR', 'NCHAR'] and col_length:
                    try:
                        length_val = int(col_length)
                        if length_val <= 0 or length_val > 8000:
                            return jsonify({'success': False, 'message': f'Panjang kolom "{col_name}" harus antara 1-8000'})
                        col_def += f"({length_val})"
                    except ValueError:
                        return jsonify({'success': False, 'message': f'Panjang kolom "{col_name}" harus berupa angka'})
                elif col_type in ['DECIMAL', 'NUMERIC'] and col_length:
                    # For decimal types, allow precision,scale format
                    if ',' in col_length:
                        try:
                            precision, scale = col_length.split(',')
                            precision = int(precision.strip())
                            scale = int(scale.strip())
                            if precision < 1 or precision > 38 or scale < 0 or scale > precision:
                                return jsonify({'success': False, 'message': f'Precision/Scale tidak valid untuk kolom "{col_name}"'})
                            col_def += f"({precision},{scale})"
                        except ValueError:
                            return jsonify({'success': False, 'message': f'Format precision,scale tidak valid untuk kolom "{col_name}"'})
                    else:
                        try:
                            precision = int(col_length)
                            if precision < 1 or precision > 38:
                                return jsonify({'success': False, 'message': f'Precision tidak valid untuk kolom "{col_name}"'})
                            col_def += f"({precision})"
                        except ValueError:
                            return jsonify({'success': False, 'message': f'Precision harus berupa angka untuk kolom "{col_name}"'})
                
                # Add NULL/NOT NULL
                if not allow_nulls:
                    col_def += " NOT NULL"
                else:
                    col_def += " NULL"
                
                # BAGIAN PERBAIKAN: Handle default values dengan validasi yang lebih ketat
                if default_value not in [None, '']:
                    if col_type.upper() in ['VARCHAR', 'NVARCHAR', 'CHAR', 'NCHAR', 'TEXT']:
                        # Properly escape single quotes and wrap with quotes
                        print(f"Default value for {col_name}: {default_value}")
                        escaped_value = default_value.replace("'", "''")
                        print(f"Escaped value for {col_name}: {escaped_value}")
                        col_def += f" DEFAULT '{escaped_value}'"
                    elif col_type.upper() in ['BIT']:
                        if default_value.lower() in ['0', '1', 'false', 'true']:
                            bit_value = '1' if default_value.lower() in ['1', 'true'] else '0'
                            col_def += f" DEFAULT {bit_value}"
                        else:
                            return jsonify({'success': False, 'message': f'Default value untuk kolom "{col_name}" bertipe BIT harus 0, 1, true, atau false'})
                    elif col_type.upper() in ['DATE', 'DATETIME', 'DATETIME2']:
                        # Handle special datetime functions
                        if default_value.upper() in ['GETDATE()', 'GETUTCDATE()', 'SYSDATETIME()', 'CURRENT_TIMESTAMP']:
                            col_def += f" DEFAULT {default_value.upper()}"
                        else:
                            # Validate date format
                            try:
                                from datetime import datetime
                                # Try multiple date formats
                                date_formats = ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%m/%d/%Y', '%d/%m/%Y']
                                parsed_date = None
                                for fmt in date_formats:
                                    try:
                                        parsed_date = datetime.strptime(default_value, fmt)
                                        break
                                    except ValueError:
                                        continue
                                
                                if parsed_date is None:
                                    raise ValueError("Invalid date format")
                                    
                                # Format untuk SQL Server
                                formatted_date = parsed_date.strftime('%Y-%m-%d')
                                col_def += f" DEFAULT '{formatted_date}'"
                            except ValueError:
                                return jsonify({'success': False, 'message': f'Default value untuk kolom "{col_name}" harus berformat YYYY-MM-DD atau fungsi seperti GETDATE()'})
                    elif col_type.upper() in ['TIME']:
                        # Handle time format
                        try:
                            from datetime import datetime
                            time_formats = ['%H:%M:%S', '%H:%M', '%I:%M:%S %p', '%I:%M %p']
                            parsed_time = None
                            for fmt in time_formats:
                                try:
                                    parsed_time = datetime.strptime(default_value, fmt)
                                    break
                                except ValueError:
                                    continue
                            
                            if parsed_time is None:
                                raise ValueError("Invalid time format")
                                
                            formatted_time = parsed_time.strftime('%H:%M:%S')
                            col_def += f" DEFAULT '{formatted_time}'"
                        except ValueError:
                            return jsonify({'success': False, 'message': f'Default value untuk kolom "{col_name}" harus berformat HH:MM:SS'})
                    elif col_type.upper() in ['INT', 'BIGINT', 'SMALLINT', 'TINYINT', 'DECIMAL', 'NUMERIC', 'FLOAT', 'REAL']:
                        # Numeric types - validate that it's a number
                        try:
                            # Additional validation for integer ranges
                            if col_type.upper() == 'TINYINT':
                                val = int(float(default_value))
                                if val < 0 or val > 255:
                                    return jsonify({'success': False, 'message': f'Default value untuk kolom "{col_name}" bertipe TINYINT harus antara 0-255'})
                            elif col_type.upper() == 'SMALLINT':
                                val = int(float(default_value))
                                if val < -32768 or val > 32767:
                                    return jsonify({'success': False, 'message': f'Default value untuk kolom "{col_name}" bertipe SMALLINT harus antara -32768 hingga 32767'})
                            elif col_type.upper() == 'INT':
                                val = int(float(default_value))
                                if val < -2147483648 or val > 2147483647:
                                    return jsonify({'success': False, 'message': f'Default value untuk kolom "{col_name}" bertipe INT harus antara -2147483648 hingga 2147483647'})
                            
                            # Validate float for decimal types
                            float(default_value)
                            col_def += f" DEFAULT {default_value}"
                        except ValueError:
                            return jsonify({'success': False, 'message': f'Default value untuk kolom "{col_name}" harus berupa angka yang valid'})
                        except OverflowError:
                            return jsonify({'success': False, 'message': f'Default value untuk kolom "{col_name}" terlalu besar'})
                    else:
                        # For other types, wrap in quotes as string
                        escaped_value = default_value.replace("'", "''")
                        col_def += f" DEFAULT '{escaped_value}'"
                elif not allow_nulls:
                    # Jika tidak allow nulls dan tidak ada default value, berikan error
                    return jsonify({'success': False, 'message': f'Kolom "{col_name}" harus memiliki default value karena tidak mengizinkan NULL'})

                column_definitions.append(col_def)
            
            # Add automatic columns
            column_definitions.append("    [period_date] DATE NULL")
            column_definitions.append("    [upload_date] DATETIME NOT NULL DEFAULT GETDATE()")
            
            create_query += ",\n".join(column_definitions)
            create_query += "\n)"
            
            logger.info(f"Creating table with query: {create_query}")
            
            # Execute dan verify dengan handling yang lebih baik
            try:
                # Insert to MasterCreator first
                cursor.execute("""
                    INSERT INTO MasterCreator (template_name, division_name, create_date, create_by)
                    VALUES (?, ?, GETDATE(), ?)
                """, (table_name, divisions, username))
                
                # Create the table
                cursor.execute(create_query)
                
                # Wait and verify table creation
                import time
                time.sleep(0.5)
                
                # Verify table was created successfully
                cursor.execute("""
                    SELECT COUNT(*) FROM sys.tables 
                    WHERE name = ? AND type = 'U'
                """, (table_name,))
                
                table_count = cursor.fetchone()[0]
                
                if table_count == 0:
                    conn.rollback()
                    logger.error(f"Template {table_name} was not created successfully")
                    return jsonify({'success': False, 'message': 'Tabel gagal dibuat. Silakan periksa log database.'})
                
                # Verify columns are created correctly
                cursor.execute("""
                    SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_NAME = ?
                """, (table_name,))
                
                column_count = cursor.fetchone()[0]
                expected_count = len(columns) + 3  # +3 for id, period_date, upload_date
                
                if column_count != expected_count:
                    logger.warning(f"Template {table_name} created but column count mismatch. Expected: {expected_count}, Actual: {column_count}")
                
                # Commit hanya jika verifikasi berhasil
                conn.commit()
                logger.info(f"Template {table_name} created and verified successfully with {column_count} columns")
                
                return jsonify({
                    'success': True,
                    'message': f'Template "{table_name}" berhasil dibuat dengan {column_count} kolom.',
                    'table_name': table_name,
                    'columns_created': len(columns),
                    'total_columns': column_count,
                    'automatic_columns': ['id', 'period_date', 'upload_date'],
                    'query': create_query
                })
                
            except Exception as create_error:
                conn.rollback()
                logger.error(f"Error creating table {table_name}: {str(create_error)}")
                
                # Handle specific SQL Server errors
                error_msg = str(create_error)
                if "42S01" in error_msg or "2714" in error_msg:
                    return jsonify({'success': False, 'message': f'Template "{table_name}" sudah ada dalam database'})
                elif "2705" in error_msg:
                    return jsonify({'success': False, 'message': 'Nama kolom duplikat atau tidak valid'})
                elif "102" in error_msg:
                    return jsonify({'success': False, 'message': 'Syntax error dalam query SQL. Periksa tipe data dan default values.'})
                elif "245" in error_msg:
                    return jsonify({'success': False, 'message': 'Error konversi tipe data. Periksa default values yang dimasukkan.'})
                elif "2627" in error_msg:
                    return jsonify({'success': False, 'message': 'Terdapat duplikasi data atau constraint violation.'})
                else:
                    return jsonify({'success': False, 'message': f'Error database: {error_msg}'})
            
        except Exception as e:
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
            logger.error(f"Error creating table: {str(e)}")
            return jsonify({'success': False, 'message': f'Error: {str(e)}'})
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
                                
@app.route('/get-template-details/<template_name>')
def get_template_details(template_name):
    """
    Get detailed information about a template including columns
    """
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first'})
    
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get template basic info from MasterCreator
        cursor.execute("""
            SELECT mc.template_name, mc.division_name, mc.create_date, mc.create_by
            FROM MasterCreator mc
            WHERE mc.template_name = ?
        """, (template_name,))
        
        template_info = cursor.fetchone()
        
        if not template_info:
            return jsonify({'success': False, 'message': 'Template not found in MasterCreator'})
        
        # Get column information from INFORMATION_SCHEMA
        cursor.execute("""
            SELECT 
                c.COLUMN_NAME,
                c.DATA_TYPE,
                c.CHARACTER_MAXIMUM_LENGTH,
                c.NUMERIC_PRECISION,
                c.NUMERIC_SCALE,
                c.IS_NULLABLE,
                c.COLUMN_DEFAULT,
                c.ORDINAL_POSITION,
                CASE WHEN pk.COLUMN_NAME IS NOT NULL THEN 1 ELSE 0 END as IS_PRIMARY_KEY,
                CASE WHEN COLUMNPROPERTY(OBJECT_ID(c.TABLE_SCHEMA + '.' + c.TABLE_NAME), c.COLUMN_NAME, 'IsIdentity') = 1 
                     THEN 1 ELSE 0 END as IS_IDENTITY
            FROM INFORMATION_SCHEMA.COLUMNS c
            LEFT JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE pk ON 
                c.TABLE_NAME = pk.TABLE_NAME AND 
                c.COLUMN_NAME = pk.COLUMN_NAME AND 
                pk.CONSTRAINT_NAME LIKE 'PK_%'
            WHERE c.TABLE_NAME = ?
            ORDER BY c.ORDINAL_POSITION
        """, (template_name,))
        
        columns_raw = cursor.fetchall()
        
        if not columns_raw:
            return jsonify({'success': False, 'message': f'Template "{template_name}" not found in database'})
        
        # Format column information
        columns = []
        for col in columns_raw:
            # Format data type display
            data_type_display = col[1]
            if col[2] and col[2] != -1:  # CHARACTER_MAXIMUM_LENGTH
                data_type_display += f"({col[2]})"
            elif col[3] and col[4] is not None:  # NUMERIC_PRECISION and SCALE
                data_type_display += f"({col[3]},{col[4]})"
            elif col[3]:  # NUMERIC_PRECISION only
                data_type_display += f"({col[3]})"
            
            # Clean up default value display
            default_display = col[6]
            if default_display:
                # Remove extra parentheses from SQL Server default values
                if default_display.startswith('(') and default_display.endswith(')'):
                    default_display = default_display[1:-1]
                # Remove quotes from string defaults
                if default_display.startswith("'") and default_display.endswith("'"):
                    default_display = default_display[1:-1]
            
            column_info = {
                'name': col[0],
                'data_type': data_type_display,
                'raw_data_type': col[1],
                'max_length': col[2],
                'numeric_precision': col[3],
                'numeric_scale': col[4],
                'is_nullable': col[5] == 'YES',
                'default_value': default_display,
                'raw_default_value': col[6],
                'ordinal_position': col[7],
                'is_primary_key': bool(col[8]),
                'is_identity': bool(col[9])
            }
            columns.append(column_info)
        
        template_details = {
            'name': template_info[0],
            'division': template_info[1],
            'create_date': template_info[2].strftime('%Y-%m-%d %H:%M:%S') if template_info[2] else None,
            'created_by': template_info[3],
            'total_columns': len(columns),
            'user_columns': len([c for c in columns if c['name'] not in ['id', 'period_date', 'upload_date']]),
            'columns': columns
        }
        
        return jsonify({
            'success': True,
            'template': template_details
        })
        
    except Exception as e:
        logger.error(f"Error getting template details for {template_name}: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/delete-table', methods=['POST'])
def delete_table():
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})
    
    try:
        data = request.get_json()
        table_name = data.get('table_name', '').strip()
        
        if not table_name:
            return jsonify({'success': False, 'message': 'Table name is required'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        conn.autocommit = False
        
        try:
            # Check if it's a template first
            cursor.execute("""
                SELECT COUNT(*) FROM MasterCreator 
                WHERE template_name = ?
            """, (table_name,))
            
            is_template = cursor.fetchone()[0] > 0
            
            if is_template:
                # Delete from template table
                cursor.execute("""
                    DELETE FROM MasterCreator 
                    WHERE template_name = ?
                """, (table_name,))
                
                # Also check and drop actual table if it exists
                cursor.execute("""
                    SELECT COUNT(*) FROM sys.tables 
                    WHERE name = ? AND type = 'U'
                """, (table_name,))
                
                table_exists = cursor.fetchone()[0] > 0
                
                if table_exists:
                    cursor.execute(f"DROP TABLE [{table_name}]")
                    logger.info(f"Dropped table {table_name}")
                
                conn.commit()
                return jsonify({
                    'success': True, 
                    'message': f'Template "{table_name}" berhasil dihapus'
                })
            else:
                # Check if it's an actual table
                cursor.execute("""
                    SELECT COUNT(*) FROM sys.tables 
                    WHERE name = ? AND type = 'U'
                """, (table_name,))
                
                table_exists = cursor.fetchone()[0] > 0
                
                if table_exists:
                    cursor.execute(f"DROP TABLE [{table_name}]")
                    conn.commit()
                    logger.info(f"Dropped table {table_name}")
                    return jsonify({
                        'success': True, 
                        'message': f'Table "{table_name}" berhasil dihapus'
                    })
                else:
                    return jsonify({
                        'success': False, 
                        'message': f'Table "{table_name}" not found'
                    })
                    
        except Exception as e:
            conn.rollback()
            logger.error(f"Error deleting table {table_name}: {str(e)}")
            return jsonify({
                'success': False, 
                'message': f'Error deleting table: {str(e)}'
            })
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        logger.error(f"Error in delete_table: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# CRUD Users Management - Simplified Backend Code (Create & Delete Only)
@app.route('/users', methods=['GET', 'POST'])
def handle_users():
    if 'username' not in session:
        flash('Please log in first.')
        return redirect(url_for('login'))

    if request.method == 'GET':
        # Jika browser meminta HTML (bukan fetch/ajax)
        if request.headers.get('Accept', '').startswith('text/html'):
            return render_template(
                'users_management.html',
                username=session.get('username'),
                fullname=session.get('fullname'),
                division=session.get('division'),
                role_access=session.get('role_access')
            )

        # Jika permintaan fetch() dari JavaScript, balas JSON
        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT id, username, role_access, 
                    fullname, email, division
                FROM MasterUsers
                ORDER BY created_date DESC
            """)
            
            rows = cursor.fetchall()
            columns = [column[0] for column in cursor.description]
            users = [dict(zip(columns, row)) for row in rows]

            return jsonify({'success': True, 'users': users})

        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    elif request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        password_confirm = data.get('password_confirm')
        role_access = data.get('role_access')
        fullname = data.get('fullname')
        email = data.get('email')
        division = data.get('division')
        created_date = datetime.now()

        required_fields = [username, password, password_confirm, role_access, fullname, email, division]
        if not all(required_fields):
            return jsonify({'success': False, 'message': 'Please fill in all fields.'})

        if password != password_confirm:
            return jsonify({'success': False, 'message': 'Passwords do not match.'})

        try:
            conn = get_db_connection()
            cur = conn.cursor()

            cur.execute("SELECT id FROM MasterUsers WHERE username = ?", (username,))
            if cur.fetchone():
                return jsonify({'success': False, 'message': 'Username already exists.'})

            cur.execute("SELECT id FROM MasterUsers WHERE email = ?", (email,))
            if cur.fetchone():
                return jsonify({'success': False, 'message': 'Email is already registered.'})

            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

            cur.execute("""
                INSERT INTO MasterUsers (username, password_hash, role_access, fullname, email, division, created_date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (username, password_hash, role_access, fullname, email, division, created_date))

            conn.commit()
            return jsonify({
                'success': True, 
                'message': f'User "{username}" created successfully.'
                })

        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

        finally:
            if cur:
                cur.close()
            if conn:
                conn.close()

@app.route('/users/<int:id>', methods=['GET'])
def get_user_by_id(id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, username, fullname, email, division, role_access
            FROM MasterUsers WHERE id = ?
        """, (id,))
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'User not found'})
        user = dict(zip([desc[0] for desc in cur.description], row))
        return jsonify({'success': True, 'user': user})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        cur.close()
        conn.close()

@app.route('/users/<int:id>', methods=['PUT'])
def update_user(id):
    if 'username' not in session:
        flash('Please log in first.')
        return redirect(url_for('login'))

    try:
        data = request.get_json()

        username = data.get('username')
        fullname = data.get('fullname')
        email = data.get('email')
        division = data.get('division')
        role_access = data.get('role_access')

        if not all([username, fullname, email, division, role_access]):
            return jsonify({'success': False, 'message': 'All fields are required'})

        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if user exists
        cursor.execute("SELECT id FROM MasterUsers WHERE id = ?", (id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'User not found'})

        # Check if username is taken by another user
        cursor.execute("SELECT id FROM MasterUsers WHERE username = ? AND id != ?", (username, id))
        if cursor.fetchone():
            return jsonify({'success': False, 'message': 'Username already exists'})

        # Check if email is taken by another user
        cursor.execute("SELECT id FROM MasterUsers WHERE email = ? AND id != ?", (email, id))
        if cursor.fetchone():
            return jsonify({'success': False, 'message': 'Email is already registered'})

        cursor.execute("""
            UPDATE MasterUsers
            SET username = ?, fullname = ?, email = ?, division = ?, role_access = ?
            WHERE id = ?
        """, (username, fullname, email, division, role_access, id))

        conn.commit()
        return jsonify({
            'success': True, 
            'message': f'User "{username}" updated successfully'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/users/<int:id>', methods=['DELETE'])
def delete_user(id):
    if 'username' not in session:
        flash('Please log in first.')
        return redirect(url_for('login'))

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get division name before deletion
        cursor.execute("SELECT username FROM MasterUsers WHERE id = ?", (id,))
        result = cursor.fetchone()
        
        if not result:
            return jsonify({'success': False, 'message': 'User not found'})
        
        username = result[0]
        
        # Delete the division
        cursor.execute("DELETE FROM MasterUsers WHERE id = ?", (id,))
        conn.commit()
        
        return jsonify({'success': True, 'message': f'User: "{username}" deleted successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# Tambahkan route baru untuk validasi username
@app.route('/users/check-username', methods=['POST'])
def check_username():
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first.'})
    
    data = request.get_json()
    username = data.get('username')
    user_id = data.get('user_id')  # Optional, untuk edit user
    
    if not username:
        return jsonify({'success': False, 'message': 'Username is required'})
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if user_id:
            # Untuk edit user - cek username selain user yang sedang diedit
            cursor.execute("SELECT id FROM MasterUsers WHERE username = ? AND id != ?", (username, user_id))
        else:
            # Untuk create user baru
            cursor.execute("SELECT id FROM MasterUsers WHERE username = ?", (username,))
        
        exists = cursor.fetchone() is not None
        
        return jsonify({
            'success': True,
            'exists': exists,
            'message': 'Username already exists' if exists else 'Username available'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/users/check-email', methods=['POST'])
def check_email():
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first.'})
    
    data = request.get_json()
    email = data.get('email')
    user_id = data.get('user_id')  # Optional, untuk edit user
    
    if not email:
        return jsonify({'success': False, 'message': 'Email is required'})
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if user_id:
            # Untuk edit user - cek email selain user yang sedang diedit
            cursor.execute("SELECT id FROM MasterUsers WHERE email = ? AND id != ?", (email, user_id))
        else:
            # Untuk create user baru
            cursor.execute("SELECT id FROM MasterUsers WHERE email = ?", (email,))
        
        exists = cursor.fetchone() is not None
        
        return jsonify({
            'success': True,
            'exists': exists,
            'message': 'Email already registered' if exists else 'Email available'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()             

# CRUD Divisions Management - Simplified Backend Code (Create & Delete Only)
@app.route('/divisions-page')
def divisions_page():
    if 'username' not in session:
        flash('Please log in first.')
        return redirect(url_for('login'))

    return render_template(
        'divisions_management.html',
        username=session.get('username'),
        fullname=session.get('fullname'),
        division=session.get('division'),
        role_access=session.get('role_access')
    )

@app.route('/divisions', methods=['GET'])
def get_divisions():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, division_name, created_by, 
                   CONVERT(VARCHAR(19), created_date, 120) as created_date
            FROM MasterDivisions
            ORDER BY created_date DESC
        """)
        
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        divisions = [dict(zip(columns, row)) for row in rows]

        return jsonify({'success': True, 'divisions': divisions})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# Separate endpoint for dropdown options (simple format)
@app.route('/divisions/dropdown', methods=['GET'])
def get_divisions_dropdown():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT division_name FROM MasterDivisions WHERE division_name IS NOT NULL ORDER BY division_name")
        rows = cursor.fetchall()
        divisions = [r[0] for r in rows]
        
        return jsonify({'success': True, 'divisions': divisions})
        
    except Exception as e:
        print(f"Error in get_divisions_dropdown: {str(e)}")
        return jsonify({'success': False, 'message': f'Failed to load divisions: {str(e)}'})
    
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/divisions', methods=['POST'])
def create_division():
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first'})

    conn = None
    cursor = None
    
    try:
        data = request.get_json()
        division_name = data.get('division_name', '').strip()
        created_by = session.get('username')

        if not division_name:
            return jsonify({'success': False, 'message': 'Nama divisi harus diisi'})

        # Validate division name format
        if not re.match(r'^[a-zA-Z0-9\s_-]+$', division_name):
            return jsonify({'success': False, 'message': 'Nama divisi hanya boleh mengandung huruf, angka, spasi, underscore, dan dash'})

        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if division name already exists (case insensitive)
        cursor.execute("""
            SELECT COUNT(*) FROM MasterDivisions 
            WHERE LOWER(division_name) = LOWER(?)
        """, (division_name,))
        
        if cursor.fetchone()[0] > 0:
            return jsonify({'success': False, 'message': f'Nama divisi "{division_name}" sudah ada'})

        # Insert new division
        cursor.execute("""
            INSERT INTO MasterDivisions (division_name, created_by, created_date)
            VALUES (?, ?, GETDATE())
        """, (division_name, created_by))
        
        conn.commit()
        
        # Log successful creation
        logger.info(f"Division '{division_name}' created successfully by {created_by}")
        
        return jsonify({
            'success': True, 
            'message': f'Divisi "{division_name}" berhasil dibuat',
            'division_name': division_name
        })
        
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        logger.error(f"Error creating division: {str(e)}")
        return jsonify({'success': False, 'message': f'Terjadi kesalahan: {str(e)}'})
        
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/divisions/<int:division_id>', methods=['DELETE'])
def delete_division(division_id):
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first'})

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get division name before deletion
        cursor.execute("SELECT division_name FROM MasterDivisions WHERE id = ?", (division_id,))
        result = cursor.fetchone()
        
        if not result:
            return jsonify({'success': False, 'message': 'Division not found'})
        
        division_name = result[0]
        
        # Delete the division
        cursor.execute("DELETE FROM MasterDivisions WHERE id = ?", (division_id,))
        conn.commit()
        
        return jsonify({'success': True, 'message': f'Division "{division_name}" deleted successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
   
@app.route('/check-period', methods=['POST'])
def check_period():
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first.'})

    data = request.json
    table_name = data.get('table_name')
    periode_date = data.get('periode_date')

    if not table_name or not periode_date:
        return jsonify({'success': False, 'message': 'Nama tabel dan periode_date wajib diisi.'})

    try:
        # 🔹 Ubah dari YYYY-MM menjadi YYYY-MM-01
        try:
            periode_date = datetime.strptime(periode_date, '%Y-%m').date().replace(day=1)
        except ValueError:
            return jsonify({'success': False, 'message': 'Format tanggal periode tidak valid. Gunakan format YYYY-MM'})

        conn = get_db_connection()
        cursor = conn.cursor()
        query = f"SELECT COUNT(*) FROM {table_name} WHERE period_date = ?"
        cursor.execute(query, (periode_date,))
        count = cursor.fetchone()[0]
        return jsonify({'success': True, 'exists': count > 0})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error checking period: {str(e)}'})

    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# PERBAIKAN 2: Endpoint check table yang lebih reliable
@app.route('/check-table-exists/<table_name>')
def check_table_exists(table_name):
    """Check if table already exists in database"""
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # PERBAIKAN: Gunakan multiple check untuk memastikan
        cursor.execute("""
            SELECT COUNT(*) FROM sys.tables 
            WHERE name = ? AND type = 'U'
        """, (table_name,))
        
        exists_count = cursor.fetchone()[0]
        exists = exists_count > 0
        
        # PERBAIKAN: Double check dengan INFORMATION_SCHEMA
        cursor.execute("""
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_NAME = ? AND TABLE_TYPE = 'BASE TABLE'
        """, (table_name,))
        
        exists_info_schema = cursor.fetchone()[0] > 0
        
        # Jika ada perbedaan, log warning
        if exists != exists_info_schema:
            logger.warning(f"Table existence check mismatch for {table_name}: sys.tables={exists}, INFORMATION_SCHEMA={exists_info_schema}")
        
        # Gunakan hasil yang paling konservatif (jika salah satu mengatakan ada, maka ada)
        final_exists = exists or exists_info_schema
        
        return jsonify({
            'success': True,
            'exists': final_exists,
            'message': f'Tabel "{table_name}" {"sudah ada" if final_exists else "belum ada"} dalam database'
        })
        
    except Exception as e:
        logger.error(f"Error checking table existence: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/get-existing-tables')
def get_existing_tables():
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})

    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Get user role and division for filtering
        role_access = session.get('role_access')
        user_division = session.get('division')

        # Build query based on role
        if role_access == 'admin':
            # Admin can see all templates
            query = """
                SELECT 
                    template_name,
                    division_name,
                    create_date,
                    create_by
                FROM MasterCreator
                ORDER BY create_date DESC
            """
            cursor.execute(query)
        else:
            # Non-admin users can only see their division's templates
            query = """
                SELECT 
                    template_name,
                    division_name,
                    create_date,
                    create_by
                FROM MasterCreator
                WHERE division_name = ?
                ORDER BY create_date DESC
            """
            cursor.execute(query, (user_division,))

        templates = cursor.fetchall()
        
        # Log for debugging
        logger.info(f"Found {len(templates)} templates for user {session.get('username')}")

        # Build response
        tables = []
        for template in templates:
            table_data = {
                'name': template[0],
                'division': template[1],
                'create_date': template[2].strftime('%Y-%m-%d %H:%M:%S') if template[2] else 'N/A',
                'create_by': template[3]
            }
            tables.append(table_data)
            logger.debug(f"Added template: {table_data}")

        return jsonify({
            'success': True, 
            'tables': tables,
            'count': len(tables)
        })

    except Exception as e:
        logger.error(f"Error getting existing tables: {str(e)}")
        return jsonify({
            'success': False, 
            'message': f'Database error: {str(e)}'
        })
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/get-table-details/<table_name>', methods=['GET'])
def get_table_details(table_name):
    """Get detailed information about a specific table"""
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first'})
    
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get table columns information
        cursor.execute("""
            SELECT 
                c.name as column_name,
                t.name as data_type,
                c.max_length,
                c.precision,
                c.scale,
                c.is_nullable,
                c.is_identity
            FROM sys.columns c
            INNER JOIN sys.types t ON c.user_type_id = t.user_type_id
            INNER JOIN sys.tables tb ON c.object_id = tb.object_id
            WHERE tb.name = ? AND tb.type = 'U'
            ORDER BY c.column_id
        """, (table_name,))
        
        columns = []
        for row in cursor.fetchall():
            column_name = row[0]
            data_type = row[1]
            max_length = row[2]
            precision = row[3]
            scale = row[4]
            is_nullable = row[5]
            is_identity = row[6]
            
            # Skip identity columns (like ID)
            if is_identity:
                continue
                
            # Format length based on data type
            length = ''
            if data_type in ['varchar', 'nvarchar', 'char', 'nchar'] and max_length > 0:
                length = str(max_length if max_length != -1 else 'MAX')
            elif data_type in ['decimal', 'numeric'] and precision > 0:
                length = f"{precision},{scale}" if scale > 0 else str(precision)
            
            columns.append({
                'name': column_name,
                'type': data_type.upper(),
                'length': length,
                'nullable': is_nullable
            })
        
        return jsonify({
            'success': True,
            'table': {
                'name': table_name,
                'columns': columns
            }
        })
        
    except Exception as e:
        logger.error(f"Error fetching table details for {table_name}: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching table details: {str(e)}'
        })
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/get-table-data/<table_name>', methods=['GET'])
def get_table_data(table_name):
    """Get data from a specific table"""
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first'})
    
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # First, check if table exists
        cursor.execute("""
            SELECT COUNT(*) FROM sys.tables 
            WHERE name = ? AND type = 'U'
        """, (table_name,))
        
        if cursor.fetchone()[0] == 0:
            return jsonify({'success': False, 'message': f'Table "{table_name}" not found'})
        
        # Get table data with pagination
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        offset = (page - 1) * per_page
        
        # Get total count
        cursor.execute(f"SELECT COUNT(*) FROM [{table_name}]")
        total_count = cursor.fetchone()[0]
        
        # Get data with pagination
        cursor.execute(f"""
            SELECT * FROM [{table_name}]
            ORDER BY id DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """, (offset, per_page))
        
        columns = [column[0] for column in cursor.description]
        data = []
        
        for row in cursor.fetchall():
            row_data = {}
            for i, column in enumerate(columns):
                value = row[i]
                if isinstance(value, datetime):
                    row_data[column] = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, date):
                    row_data[column] = value.strftime('%Y-%m-%d')
                else:
                    row_data[column] = value
            data.append(row_data)
        
        return jsonify({
            'success': True,
            'data': data,
            'columns': columns,
            'total': total_count,
            'page': page,
            'per_page': per_page,
            'total_pages': (total_count + per_page - 1) // per_page
        })
        
    except Exception as e:
        logger.error(f"Error fetching table data for {table_name}: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching table data: {str(e)}'
        })
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
                                    
@app.route('/get-sql-data-types')
def get_sql_data_types():
    """Get available SQL Server data types"""
    data_types = [
        {'value': 'VARCHAR', 'label': 'VARCHAR - Variable-length character string', 'has_length': True},
        {'value': 'NVARCHAR', 'label': 'NVARCHAR - Variable-length Unicode string', 'has_length': True},
        {'value': 'CHAR', 'label': 'CHAR - Fixed-length character string', 'has_length': True},
        {'value': 'NCHAR', 'label': 'NCHAR - Fixed-length Unicode string', 'has_length': True},
        {'value': 'TEXT', 'label': 'TEXT - Variable-length text data', 'has_length': False},
        {'value': 'NTEXT', 'label': 'NTEXT - Variable-length Unicode text', 'has_length': False},
        {'value': 'INT', 'label': 'INT - Integer (32-bit)', 'has_length': False},
        {'value': 'BIGINT', 'label': 'BIGINT - Large integer (64-bit)', 'has_length': False},
        {'value': 'SMALLINT', 'label': 'SMALLINT - Small integer (16-bit)', 'has_length': False},
        {'value': 'TINYINT', 'label': 'TINYINT - Very small integer (8-bit)', 'has_length': False},
        {'value': 'DECIMAL', 'label': 'DECIMAL - Exact numeric with precision and scale', 'has_length': True},
        {'value': 'NUMERIC', 'label': 'NUMERIC - Exact numeric with precision and scale', 'has_length': True},
        {'value': 'FLOAT', 'label': 'FLOAT - Approximate numeric floating point', 'has_length': False},
        {'value': 'REAL', 'label': 'REAL - Approximate numeric floating point', 'has_length': False},
        {'value': 'MONEY', 'label': 'MONEY - Monetary data', 'has_length': False},
        {'value': 'SMALLMONEY', 'label': 'SMALLMONEY - Monetary data (smaller range)', 'has_length': False},
        {'value': 'DATE', 'label': 'DATE - Date only', 'has_length': False},
        {'value': 'TIME', 'label': 'TIME - Time only', 'has_length': False},
        {'value': 'DATETIME', 'label': 'DATETIME - Date and time', 'has_length': False},
        {'value': 'DATETIME2', 'label': 'DATETIME2 - Date and time with higher precision', 'has_length': False},
        {'value': 'SMALLDATETIME', 'label': 'SMALLDATETIME - Date and time (smaller range)', 'has_length': False},
        {'value': 'BIT', 'label': 'BIT - Boolean (0 or 1)', 'has_length': False},
        {'value': 'UNIQUEIDENTIFIER', 'label': 'UNIQUEIDENTIFIER - GUID', 'has_length': False},
        {'value': 'VARBINARY', 'label': 'VARBINARY - Variable-length binary data', 'has_length': True},
        {'value': 'BINARY', 'label': 'BINARY - Fixed-length binary data', 'has_length': True},
        {'value': 'IMAGE', 'label': 'IMAGE - Variable-length binary data (legacy)', 'has_length': False}
    ]
    
    return jsonify({
        'success': True,
        'data_types': data_types
    })

@app.route('/get-excel-sheets', methods=['POST'])
def get_excel_sheets_endpoint():
    """Endpoint untuk mendapatkan daftar sheet dalam file Excel"""
    
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first.'})
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Tidak ada file yang dipilih'})
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Tidak ada file yang dipilih'})
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': 'File harus berformat Excel (.xlsx atau .xls)'})
        
        # Simpan file temporary
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"sheets_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        try:
            # Dapatkan daftar sheet
            sheets = get_excel_sheets(file_path)
            
            if sheets is None:
                return jsonify({'success': False, 'message': 'Gagal membaca daftar sheet dari file Excel'})
            
            return jsonify({
                'success': True,
                'sheets': sheets,
                'message': f'Ditemukan {len(sheets)} sheet dalam file Excel'
            })
            
        finally:
            # Hapus file temporary
            try:
                os.remove(file_path)
            except:
                pass
        
    except Exception as e:
        logger.error(f"Error in get_excel_sheets: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/export-table/<table_name>', methods=['GET'])
def export_table(table_name):
    """Export table data to CSV"""
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first'})
    
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if table exists
        cursor.execute("""
            SELECT COUNT(*) FROM sys.tables 
            WHERE name = ? AND type = 'U'
        """, (table_name,))
        
        if cursor.fetchone()[0] == 0:
            return jsonify({'success': False, 'message': f'Table "{table_name}" not found'})
        
        # Get all data
        cursor.execute(f"SELECT * FROM [{table_name}] ORDER BY id")
        
        # Create CSV response
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        columns = [column[0] for column in cursor.description]
        writer.writerow(columns)
        
        # Write data
        for row in cursor.fetchall():
            processed_row = []
            for value in row:
                if isinstance(value, datetime):
                    processed_row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                elif isinstance(value, date):
                    processed_row.append(value.strftime('%Y-%m-%d'))
                else:
                    processed_row.append(value)
            writer.writerow(processed_row)
        
        # Prepare response
        output.seek(0)
        csv_content = output.getvalue()
        
        response = make_response(csv_content)
        response.headers["Content-Disposition"] = f"attachment; filename={table_name}_export.csv"
        response.headers["Content-Type"] = "text/csv"
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting table {table_name}: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error exporting table: {str(e)}'
        })
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/duplicate-table/<table_name>', methods=['POST'])
def duplicate_table(table_name):
    """Duplicate an existing table structure"""
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first'})
    
    conn = None
    cursor = None
    try:
        data = request.get_json()
        new_table_name = data.get('new_table_name', '').strip()
        
        if not new_table_name:
            return jsonify({'success': False, 'message': 'New table name is required'})
        
        if not re.match(r'^[a-zA-Z][a-zA-Z0-9_]*$', new_table_name):
            return jsonify({'success': False, 'message': 'Invalid table name format'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        conn.autocommit = False
        
        # Check if source table exists
        cursor.execute("""
            SELECT COUNT(*) FROM sys.tables 
            WHERE name = ? AND type = 'U'
        """, (table_name,))
        
        if cursor.fetchone()[0] == 0:
            return jsonify({'success': False, 'message': f'Source table "{table_name}" not found'})
        
        # Check if new table name already exists
        cursor.execute("""
            SELECT COUNT(*) FROM sys.tables 
            WHERE name = ? AND type = 'U'
        """, (new_table_name,))
        
        if cursor.fetchone()[0] > 0:
            return jsonify({'success': False, 'message': f'Table "{new_table_name}" already exists'})
        
        # Get source table structure
        cursor.execute(f"""
            SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, 
                   NUMERIC_PRECISION, NUMERIC_SCALE, IS_NULLABLE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = ? AND COLUMN_NAME NOT IN ('id', 'period_date', 'upload_date')
            ORDER BY ORDINAL_POSITION
        """, (table_name,))
        
        columns = cursor.fetchall()
        
        if not columns:
            return jsonify({'success': False, 'message': 'No columns found in source table'})
        
        # Create new table with same structure
        create_query = f"CREATE TABLE [{new_table_name}] (\n"
        create_query += "    [id] INT IDENTITY(1,1) PRIMARY KEY,\n"
        
        column_definitions = []
        for col in columns:
            col_name = col[0]
            col_type = col[1].upper()
            max_length = col[2]
            precision = col[3]
            scale = col[4]
            is_nullable = col[5]
            
            col_def = f"    [{col_name}] {col_type}"
            
            # Add length/precision
            if col_type in ['VARCHAR', 'NVARCHAR', 'CHAR', 'NCHAR'] and max_length:
                col_def += f"({max_length})"
            elif col_type in ['DECIMAL', 'NUMERIC'] and precision:
                if scale and scale > 0:
                    col_def += f"({precision},{scale})"
                else:
                    col_def += f"({precision})"
            
            col_def += " NULL" if is_nullable == 'YES' else " NOT NULL"
            column_definitions.append(col_def)
        
        # Add standard columns
        column_definitions.append("    [period_date] DATE NULL")
        column_definitions.append("    [upload_date] DATETIME NOT NULL DEFAULT GETDATE()")
        
        create_query += ",\n".join(column_definitions)
        create_query += "\n)"
        
        cursor.execute(create_query)
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': f'Table "{new_table_name}" created successfully as duplicate of "{table_name}"',
            'new_table_name': new_table_name
        })
        
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        logger.error(f"Error duplicating table: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error duplicating table: {str(e)}'
        })
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/save-as-template', methods=['POST'])
def save_as_template():
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})
    
    try:
        data = request.get_json()
        table_name = data.get('table_name', '').strip()
        columns = data.get('columns', [])
        division = data.get('division', '').strip()
        username = session.get('username')
        
        if not table_name or not columns or not division:
            return jsonify({'success': False, 'message': 'Missing required fields'})
        
        # Validate table name
        if not re.match(r'^[a-zA-Z][a-zA-Z0-9_]*$', table_name):
            return jsonify({'success': False, 'message': 'Invalid table name format'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        conn.autocommit = False
        
        try:
            # Check if template already exists
            cursor.execute("""
                SELECT COUNT(*) FROM MasterCreator 
                WHERE template_name = ?
            """, (table_name,))
            
            if cursor.fetchone()[0] > 0:
                return jsonify({'success': False, 'message': 'Template already exists'})
            
            # Save template metadata
            cursor.execute("""
                INSERT INTO MasterCreator (template_name, division_name, create_date, create_by)
                VALUES (?, ?, GETDATE(), ?)
            """, (table_name, division, username))
            
            # You might want to save column definitions in a separate table
            # For now, just save the template record
            
            conn.commit()
            logger.info(f"Template {table_name} saved successfully")
            
            return jsonify({
                'success': True,
                'message': f'Template "{table_name}" saved successfully'
            })
            
        except Exception as e:
            conn.rollback()
            logger.error(f"Error saving template {table_name}: {str(e)}")
            return jsonify({'success': False, 'message': f'Error saving template: {str(e)}'})
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        logger.error(f"Error in save_as_template: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/debitur-aktif', methods=['GET'])
def api_debitur_aktif():
    """
    GET /api/debitur-aktif
    Mengambil data debitur aktif untuk preview (max 1000 rows)
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Default: periode terakhir (EOD_DATE)
        cursor.execute("""
            SELECT MAX(PBK_EOD_DATE) FROM [10.10.4.12].SMIDWHARIUM.dbo.PBK_T_INF_COR_FACILITY_ACCOUNT
            WHERE FACILITY_STATUS = 'AC'
        """)
        last_eod_date = cursor.fetchone()[0]

        # Query data dari function
        cursor.execute("""
            SELECT TOP 1000
                PBK_EOD_DATE,
                NOMOR_CIF,
                CUST_NAME,
                FACILITY_NO
            FROM Func_GetCustomerData(?)
            ORDER BY NOMOR_CIF
        """, (last_eod_date,))
        rows = cursor.fetchall()

        # Format data
        data = []
        for idx, row in enumerate(rows):
            data.append({
                'pbk_eod_date': row[0].strftime('%Y-%m-%d') if row[0] else None,
                'kode_debitur': row[1],
                'nama_debitur': row[2],
                'facility_no': row[3],
                'status': 'AKTIF',
                'tanggal_dibuat': row[0].strftime('%Y-%m-%d') if row[0] else None,
                'last_update': row[0].strftime('%Y-%m-%d') if row[0] else None,
            })

        # Stats
        cursor.execute("""
            SELECT COUNT(DISTINCT NOMOR_CIF)
            FROM Func_GetCustomerData(?)
        """, (last_eod_date,))
        total_records = cursor.fetchone()[0]

        stats = {
            'total': total_records,
            'active': total_records,
            'last_update': last_eod_date.strftime('%Y-%m-%d') if last_eod_date else None
        }

        return jsonify({'success': True, 'data': data, 'stats': stats})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@app.route('/api/sync-debitur', methods=['POST'])
def api_sync_debitur():
    """
    POST /api/sync-debitur
    Refresh data debitur aktif dari database
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Ambil periode terbaru
        cursor.execute("""
            SELECT MAX(PBK_EOD_DATE) FROM [10.10.4.12].SMIDWHARIUM.dbo.PBK_T_INF_COR_FACILITY_ACCOUNT
            WHERE FACILITY_STATUS = 'AC'
        """)
        last_eod_date = cursor.fetchone()[0]

        # Query data dari function
        cursor.execute("""
            SELECT
                PBK_EOD_DATE,
                NOMOR_CIF,
                CUST_NAME,
                FACILITY_NO
            FROM Func_GetCustomerData(?)
            ORDER BY NOMOR_CIF
        """, (last_eod_date,))
        rows = cursor.fetchall()

        data = []
        for idx, row in enumerate(rows):
            data.append({
                'pbk_eod_date': row[0].strftime('%Y-%m-%d') if row[0] else None,
                'kode_debitur': row[1],
                'nama_debitur': row[2],
                'facility_no': row[3],
                'status': 'AKTIF',
                'tanggal_dibuat': row[0].strftime('%Y-%m-%d') if row[0] else None,
                'last_update': row[0].strftime('%Y-%m-%d') if row[0] else None,
            })

        stats = {
            'total': len(data),
            'active': len(data),
            'last_update': last_eod_date.strftime('%Y-%m-%d') if last_eod_date else None
        }

        return jsonify({'success': True, 'data': data, 'stats': stats})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@app.route('/api/download-debitur-excel', methods=['POST'])
def api_download_debitur_excel():
    """
    POST /api/download-debitur-excel
    Generate dan download file Excel debitur aktif
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Ambil periode terbaru
        cursor.execute("""
            SELECT MAX(PBK_EOD_DATE) FROM [10.10.4.12].SMIDWHARIUM.dbo.PBK_T_INF_COR_FACILITY_ACCOUNT
            WHERE FACILITY_STATUS = 'AC'
        """)
        last_eod_date = cursor.fetchone()[0]

        # Query semua data
        cursor.execute("""
            SELECT
                PBK_EOD_DATE,
                NOMOR_CIF,
                CUST_NAME,
                FACILITY_NO
            FROM Func_GetCustomerData(?)
            ORDER BY NOMOR_CIF
        """, (last_eod_date,))
        rows = cursor.fetchall()

        # Generate Excel file in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Debitur Aktif"

        # Header
        ws.append(['Tanggal Data', 'Kode Debitur', 'Nama Debitur', 'Nomor Fasilitas'])

        # Data
        for row in rows:
            ws.append([
                row[0].strftime('%Y-%m-%d') if row[0] else '',
                row[1],
                row[2],
                row[3]
            ])

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"debitur_aktif_{last_eod_date.strftime('%Y%m%d')}.xlsx" if last_eod_date else "debitur_aktif.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@app.route('/data', methods=['GET'])
def data_page():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    return render_template(
        'data.html',
        username=session.get('username'),
        fullname=session.get('fullname'),
        division=session.get('division'),
        role_access=session.get('role_access')
    )

@app.route('/api/data', methods=['GET'])
def api_data():
    """
    Endpoint untuk datatable monthly data dengan filter tanggal, pagination, dan limit
    Query params: tanggal_data, page, page_size
    """
    try:
        tanggal_data = request.args.get('tanggal_data')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 50))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Filter tanggal_data jika ada
        where_clause = ""
        params = []
        if tanggal_data:
            # Jika hanya bulan-tahun, filter dengan LIKE atau range
            if len(tanggal_data) == 7:  # format YYYY-MM
                where_clause = "WHERE CONVERT(VARCHAR(7), Tanggal_Data, 120) = ?"
                params.append(tanggal_data)
            else:
                where_clause = "WHERE Tanggal_Data = ?"
                params.append(tanggal_data)

        # Hitung total data
        count_query = f"SELECT COUNT(*) FROM [SMIDWHSSOT].[dbo].[SSOT_FINAL_MONTHLY] {where_clause}"
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()[0]

        # Ambil data dengan pagination
        offset = (page - 1) * page_size
        data_query = f"""
            SELECT *
            FROM [SMIDWHSSOT].[dbo].[SSOT_FINAL_MONTHLY]
            {where_clause}
            ORDER BY Tanggal_Data DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
        cursor.execute(data_query, params + [offset, page_size])
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        data = [dict(zip(columns, row)) for row in rows]

        return jsonify({
            'success': True,
            'data': data,
            'total': total_records,
            'page': page,
            'page_size': page_size
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
@app.route('/api/download-data', methods=['POST'])
def api_download_data():
    """
    Download data excel sesuai filter
    """
    try:
        tanggal_data = request.json.get('tanggal_data')
        conn = get_db_connection()
        cursor = conn.cursor()

        where_clause = ""
        params = []
        if tanggal_data:
            if len(tanggal_data) == 7:  # format YYYY-MM
                where_clause = "WHERE CONVERT(VARCHAR(7), Tanggal_Data, 120) = ?"
                params.append(tanggal_data)
            else:
                where_clause = "WHERE Tanggal_Data = ?"
                params.append(tanggal_data)

        # --- Ambil data utama ---
        query = f"""
            SELECT * 
            FROM [SMIDWHSSOT].[dbo].[SSOT_FINAL_MONTHLY] 
            {where_clause} 
            ORDER BY Tanggal_Data DESC
        """
        cursor.execute(query, params)
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        # --- Ambil daftar kolom numeric ---
        cursor.execute("""
            SELECT COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = 'SSOT_FINAL_MONTHLY'
            AND DATA_TYPE = 'numeric'
        """)
        numeric_columns = {row[0] for row in cursor.fetchall()}

        # --- Buat Workbook ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Data"

        # Header
        ws.append(columns)

        # Data rows
        for row in rows:
            ws.append(list(row))

        # Format kolom numeric agar tidak eksponen & bisa di-SUM
        from openpyxl.styles import numbers
        for col_idx, col_name in enumerate(columns, start=1):
            if col_name in numeric_columns:
                for row_idx in range(2, len(rows) + 2):  # skip header
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Auto adjust column width
        for col_idx, col_name in enumerate(columns, start=1):
            max_length = len(str(col_name))  # header length
            for row in rows:
                value = row[col_idx - 1]
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

        # Output ke response
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = "monthly_data.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
                        
if __name__ == '__main__':
    app.run(debug=True)