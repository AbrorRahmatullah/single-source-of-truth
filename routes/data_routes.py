from flask import (
    Blueprint, render_template, request, jsonify,
    session, redirect, url_for, send_file
)
from io import BytesIO
import openpyxl


from config.config import get_db_connection
from models.audit import insert_audit_trail

data_bp = Blueprint('data', __name__)


@data_bp.route('/data', methods=['GET'])
def data_page():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    
    insert_audit_trail('view_data_page', f"User '{session.get('username')}' accessed data page.")
    
    return render_template(
        'data.html',
        username=session.get('username'),
        fullname=session.get('fullname'),
        division=session.get('division'),
        role_access=session.get('role_access')
    )

@data_bp.route('/api/data', methods=['GET'])
def api_data():
    """
    Endpoint untuk datatable monthly data dengan filter tanggal, pagination, dan limit
    Query params: tanggal_data, page, page_size
    """
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first.'})
    
    try:
        tanggal_data = request.args.get('tanggal_data')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 50))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Filter tanggal
        where_clause = ""
        params = []
        if tanggal_data:
            if len(tanggal_data) == 7:  
                where_clause = "WHERE CONVERT(VARCHAR(7), m.Tanggal_Data, 120) = ?"
                params.append(tanggal_data)
            else:
                where_clause = "WHERE m.Tanggal_Data = ?"
                params.append(tanggal_data)

        # Hitung total
        count_query = f"SELECT COUNT(*) FROM SSOT_FINAL_MONTHLY m {where_clause}"
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()[0]

        # Ambil data + JOIN
        offset = (page - 1) * page_size
        data_query = f"""
            SELECT 
                m.*,
                ISNULL(
                    CASE 
                        WHEN m.IsSyariah = 'N' THEN 
                            CASE 
                                WHEN k.interest_reference_rate IS NULL THEN 'FIXED'
                                ELSE k.interest_reference_rate_group
                            END
                        WHEN m.IsSyariah = 'Y' THEN s.interest_reference_rate_group
                    END, 
                    m.Interest_Reference_Rate
                ) AS Interest_Reference_Rate_Group,
                
                CASE
                    WHEN m.IsSyariah = 'N' THEN 
                        CASE 
                            WHEN k.interest_reference_rate_ssot IS NULL THEN 'FIXED'
                            ELSE k.interest_reference_rate_ssot
                        END
                    WHEN m.IsSyariah = 'Y' THEN
                        CASE 
                            WHEN s.interest_reference_rate_ssot IS NULL THEN
                                ISNULL(s.interest_reference_rate_group, m.Interest_Reference_Rate)
                            ELSE s.interest_reference_rate_ssot
                        END
                END AS Interest_Reference_Rate_SSOT
                
            FROM SSOT_FINAL_MONTHLY m
            LEFT JOIN MasterInterestReferenceRateKonven k
                ON m.IsSyariah = 'N'
                AND (REPLACE(m.Interest_Reference_Rate, ' ', '') = REPLACE(k.interest_reference_rate, ' ', '')
                OR m.Interest_Reference_Rate = k.interest_reference_rate_group)
            LEFT JOIN MasterInterestReferenceRateSyariah s
                ON m.IsSyariah = 'Y'
                AND (REPLACE(m.Interest_Reference_Rate, ' ', '') = REPLACE(s.interest_reference_rate_group, ' ', '')
                OR m.Interest_Reference_Rate = s.interest_reference_rate_group)
            {where_clause}
            ORDER BY m.Tanggal_Data DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """

        cursor.execute(data_query, params + [offset, page_size])
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        data = [dict(zip(columns, row)) for row in rows]

        insert_audit_trail('view_monthly_data',
            f"User '{session.get('username')}' viewed monthly data, page {page}.")
        
        return jsonify({
            'success': True,
            'data': data,
            'total': total_records,
            'page': page,
            'page_size': page_size
        })

    except Exception as e:
        insert_audit_trail('view_monthly_data_failed',
            f"User '{session.get('username')}' failed: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
@data_bp.route('/api/download-data', methods=['POST'])
def api_download_data():
    """
    Download data excel sesuai filter
    """
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please log in first.'})
    
    try:
        tanggal_data = request.json.get('tanggal_data')
        conn = get_db_connection()
        cursor = conn.cursor()

        where_clause = ""
        params = []
        if tanggal_data:
            if len(tanggal_data) == 7:
                where_clause = "WHERE CONVERT(VARCHAR(7), m.Tanggal_Data, 120) = ?"
                params.append(tanggal_data)
            else:
                where_clause = "WHERE m.Tanggal_Data = ?"
                params.append(tanggal_data)

        # Query lengkap + JOIN
        query = f"""
            SELECT 
                m.Tanggal_Data,
                m.IsSyariah,
                m.Facility_No,
                m.Customer_Name_SLIK,
                m.Alias,
                m.Customer_ID,
                m.Kode_Revenue,
                m.Product_Name,
                m.Sub_Product_Name,
                m.Financing_Scheme,
                m.Financing_Category_Type,
                m.Nilai_Proyek,
                m.Klasifikasi_Proyek,
                m.Kategori_Proyek,
                m.Progress_Proyek,
                m.Lokasi_Proyek,
                m.Output_Proyek,
                m.Satuan_Output_Proyek,
                m.RM_PIC,
                m.Flag_Negative_Pledge,
                m.Flag_Clean_Basis,
                m.Golongan_Debitur,
                m.Nama_Kelompok_Debitur,
                m.Sektor,
                m.Sektor_Ekonomi_Lapangan_Usaha,
                m.Obyek_Pembiayaan,
                m.Kategori_Usaha_Keuangan_Berkelanjutan,
                m.Facility_Activation_Date,
                m.Perjanjian_Kredit_Date,
                m.Komitmen_ori,
                m.Komitmen_idr,
                m.Kelonggaran_Tarik,
                m.Availability_Period,
                m.OS_Principal,
                m.OS_IDR,
                m.Currency,
                m.Interest_Rate,
                m.Interest_Type,
                ISNULL(
                    CASE 
                        WHEN m.IsSyariah = 'N' THEN 
                            CASE 
                                WHEN k.interest_reference_rate IS NULL THEN 'FIXED'
                                ELSE k.interest_reference_rate_group
                            END
                        WHEN m.IsSyariah = 'Y' THEN s.interest_reference_rate_group
                    END, 
                    m.Interest_Reference_Rate
                ) AS Interest_Reference_Rate_Group,
                
                CASE
                    WHEN m.IsSyariah = 'N' THEN 
                        CASE 
                            WHEN k.interest_reference_rate_ssot IS NULL THEN 'FIXED'
                            ELSE k.interest_reference_rate_ssot
                        END
                    WHEN m.IsSyariah = 'Y' THEN
                        CASE 
                            WHEN s.interest_reference_rate_ssot IS NULL THEN
                                ISNULL(s.interest_reference_rate_group, m.Interest_Reference_Rate)
                            ELSE s.interest_reference_rate_ssot
                        END
                END AS Interest_Reference_Rate_SSOT,
                m.Maturity_Date,
                m.Start_Date_Facility,
                m.Category,
                m.Sub_Category,
                m.Divisi,
                m.Kategori_Badan_Usaha,
                m.Sub_Kategori_Badan_Usaha,
                m.Source_of_Fund,
                m.Rating_Debitur_Internal,
                m.Rating_Debitur_Eksternal,
                m.Stage,
                m.Watchlist_Flag,
                m.Flag_Restrukturisasi,
                m.Tanggal_Awal_Restru,
                m.Tanggal_Akhir_Restru,
                m.Kolektibilitas,
                m.Metode_CKPN,
                m.CKPN_Aset_Baik,
                m.CKPN_Aset_Kurang_Baik,
                m.CKPN_Aset_Tidak_Baik,
                m.CKPN,
                m.Flag_Penjaminan,
                m.Flag_Penugasan,
                m.load_date
            FROM SSOT_FINAL_MONTHLY m
            LEFT JOIN MasterInterestReferenceRateKonven k
                ON m.IsSyariah = 'N'
                AND (REPLACE(m.Interest_Reference_Rate, ' ', '') = REPLACE(k.interest_reference_rate, ' ', '')
                OR m.Interest_Reference_Rate = k.interest_reference_rate_group)
            LEFT JOIN MasterInterestReferenceRateSyariah s
                ON m.IsSyariah = 'Y'
                AND (REPLACE(m.Interest_Reference_Rate, ' ', '') = REPLACE(s.interest_reference_rate_group, ' ', '')
                OR m.Interest_Reference_Rate = s.interest_reference_rate_group)
            {where_clause}
            ORDER BY m.Tanggal_Data DESC
        """

        cursor.execute(query, params)
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        # Kolom yang dihapus
        remove_cols = {
            "Klasifikasi_Proyek",
            "Kategori_Proyek",
            "Output_Proyek",
            "Satuan_Output_Proyek"
        }

        keep_indices = [i for i, col in enumerate(columns) if col not in remove_cols]
        filtered_columns = [columns[i] for i in keep_indices]
        filtered_rows = [[row[i] for i in keep_indices] for row in rows]

        # Kolom numeric
        cursor.execute("""
            SELECT COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = 'SSOT_FINAL_MONTHLY'
            AND DATA_TYPE = 'numeric'
        """)
        numeric_columns = {row[0] for row in cursor.fetchall()}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Data"

        ws.append(filtered_columns)
        for row in filtered_rows:
            ws.append(row)

        from openpyxl.styles import numbers
        for col_idx, col_name in enumerate(filtered_columns, start=1):
            if col_name in numeric_columns:
                for row_idx in range(2, len(filtered_rows) + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Auto width
        for col_idx, col_name in enumerate(filtered_columns, start=1):
            max_len = len(str(col_name))
            for row in filtered_rows:
                val = row[col_idx - 1]
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"monthly_data_{tanggal_data}.xlsx"

        insert_audit_trail('download_monthly_data',
            f"User '{session.get('username')}' downloaded Excel.")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        insert_audit_trail('download_monthly_data_failed',
            f"User '{session.get('username')}' failed: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

    finally:
        if cursor: cursor.close()
        if conn: conn.close()
